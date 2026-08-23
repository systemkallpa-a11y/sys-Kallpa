from flask import render_template, request, jsonify, session, flash, redirect, url_for
from . import main_bp
import mysql.connector
from mysql.connector import Error
from functools import wraps
from app.config import DatabaseConfig
from .main import validar_acceso_usuario
from datetime import datetime
import openpyxl
from decimal import Decimal, InvalidOperation

def get_db_connection():
    """Crear conexin a la base de datos Kallpa"""
    try:
        params = DatabaseConfig.get_connection_params()
        connection = mysql.connector.connect(**params)
        return connection
    except Error as e:
        print(f"Error de conexin: {e}")
        return None

# Decorador para requerir autenticacin
def login_required(f):
    """Decorador para proteger rutas que requieren autenticacin"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_documento' not in session and 'user_email' not in session:
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return {'success': False, 'message': 'No autenticado'}, 401
            flash('Debes iniciar sesin', 'warning')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function

@main_bp.route('/presupuesto')
@login_required
def presupuesto():
    """Pgina principal de gestin de presupuestos"""
    num_documento = session.get('user_documento')
    
    # Validar acceso a Presupuesto
    # Menu 5 = O.T, SubMenu 9 = Presupuesto
    # Permitir si tiene acceso completo a O.T O acceso especfico a Presupuesto
    
    print(f"\n{'='*80}")
    print(f"[PRESUPUESTO_ACCESS] Validando acceso a /presupuesto")
    print(f"[PRESUPUESTO_ACCESS] Documento: {num_documento}")
    print(f"{'='*80}")
    
    # Opcin 1: Acceso completo a Menu 5 (O.T)
    print(f"[PRESUPUESTO_ACCESS] 1 Validando acceso COMPLETO a men 5 (O.T)...")
    tiene_acceso_completo = validar_acceso_usuario(num_documento, id_menu=5, id_submenu=None)
    print(f"[PRESUPUESTO_ACCESS] Resultado acceso completo: {tiene_acceso_completo}")
    
    # Opcin 2: Acceso especfico a SubMenu 9 (Presupuesto)
    print(f"[PRESUPUESTO_ACCESS] 2 Validando acceso especfico a men 5, submen 9...")
    tiene_acceso_presupuesto = validar_acceso_usuario(num_documento, id_menu=5, id_submenu=9)
    print(f"[PRESUPUESTO_ACCESS] Resultado acceso Presupuesto: {tiene_acceso_presupuesto}")
    
    print(f"[PRESUPUESTO_ACCESS] [OK] Acceso permitido: {tiene_acceso_completo or tiene_acceso_presupuesto}")
    
    if not (tiene_acceso_completo or tiene_acceso_presupuesto):
        print(f"[PRESUPUESTO_ACCESS] [X] ACCESO DENEGADO - Redirigiendo a dashboard")
        print(f"{'='*80}\n")
        flash('No tienes acceso a Gestin de Presupuestos', 'danger')
        return redirect(url_for('main.dashboard'))
    
    print(f"[PRESUPUESTO_ACCESS] [OK] ACCESO PERMITIDO - Cargando presupuesto.html")
    print(f"{'='*80}\n")
    return render_template('presupuesto.html')

@main_bp.route('/api/presupuestos/obtener', methods=['GET'])
@login_required
def obtener_presupuestos():
    """Obtener lista de presupuestos usando SP actualizado"""
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'error': 'Error de conexin'}), 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        print(f"[PRESUPUESTOS] [GET] Obteniendo reporte de presupuestos")
        
        # Intentar llamar al SP
        try:
            cursor.execute('CALL sp_ReportePresupuestos()', multi=True)
            presupuestos = cursor.fetchall()
            
            # Consumir resultados restantes si los hay
            while cursor.nextset():
                pass
        except Exception as sp_error:
            print(f"[PRESUPUESTOS] [WARN] Error con SP, usando query directa: {sp_error}")
            # Si el SP falla, usar query directa como fallback
            # IMPORTANTE: Usar GROUP BY para evitar duplicados por mltiples registros de aprobacin
            cursor.execute('''
                SELECT 
                    pr.id_presupuesto,
                    pr.numero_presupuesto,
                    pr.estado,
                    pr.monto,
                    pr.fecha_actualizacion,
                    o.nombre as nombre_obra,
                    p.nombre as nombre_proyecto,
                    CONCAT(
                        COALESCE(per.nombres, ''),
                        ' ',
                        COALESCE(per.apellido_paterno, ''),
                        ' ',
                        COALESCE(per.apellido_materno, '')
                    ) as creado_por,
                    MAX(CONCAT(
                        COALESCE(per_aprobador.nombres, ''),
                        ' ',
                        COALESCE(per_aprobador.apellido_paterno, ''),
                        ' ',
                        COALESCE(per_aprobador.apellido_materno, '')
                    )) as aprobado_rechazado_por,
                    MAX(CASE 
                        WHEN pr.estado = 'RECHAZADO' THEN COALESCE(ra.comentario, 'Sin motivo especificado')
                        ELSE NULL
                    END) as comentario_rechazo
                FROM TblPresupuesto pr
                LEFT JOIN TblObra o ON pr.id_obra = o.id_obra
                LEFT JOIN TblProyecto p ON o.id_proyecto = p.id_proyecto
                LEFT JOIN TblUsuario u ON pr.num_documento = u.num_documento
                LEFT JOIN TblPersona per ON u.num_documento = per.num_documento
                LEFT JOIN TblRegistroAprobacion ra ON 
                    pr.id_presupuesto = ra.id_documento_referencia 
                    AND ra.id_tipo_documento = 1
                    AND ra.estado_aprobacion = 'APROBADO'
                LEFT JOIN TblPersona per_aprobador ON ra.num_documento_aprobador = per_aprobador.num_documento
                WHERE pr.estado != 'ELIMINADO'
                GROUP BY pr.id_presupuesto, pr.numero_presupuesto, pr.estado, pr.monto, pr.fecha_actualizacion, o.nombre, p.nombre, per.nombres, per.apellido_paterno, per.apellido_materno
                ORDER BY pr.fecha_creacion DESC
            ''')
            presupuestos = cursor.fetchall()
        
        print(f"[PRESUPUESTOS] [OK] {len(presupuestos)} presupuestos obtenidos")
        
        cursor.close()
        connection.close()
        
        return jsonify({'success': True, 'data': presupuestos}), 200
    
    except Error as e:
        print(f"[PRESUPUESTOS] [ERROR] Error al obtener: {e}")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': str(e)}), 500
    except Exception as e:
        print(f"[PRESUPUESTOS] [ERROR] Error general: {e}")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': str(e)}), 500


@main_bp.route('/api/presupuestos/buscar', methods=['POST'])
@login_required
def buscar_presupuestos():
    """Buscar presupuestos con filtros avanzados usando SP"""
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'error': 'Error de conexin'}), 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        # Obtener filtros del request
        datos = request.get_json() or {}
        
        # Parmetros opcionales (None si no se envan)
        p_numero = datos.get('numero') or None
        p_estado = datos.get('estado') or None
        p_id_proyecto = datos.get('id_proyecto') or None
        p_id_obra = datos.get('id_obra') or None
        p_fecha_desde = datos.get('fecha_desde') or None
        p_fecha_hasta = datos.get('fecha_hasta') or None
        p_nombre_creador = datos.get('nombre_creador') or None
        p_monto_desde = datos.get('monto_desde') or None
        p_monto_hasta = datos.get('monto_hasta') or None
        
        print(f"\n{'='*80}")
        print(f"[BUSCAR_PRESUPUESTOS] Iniciando bsqueda avanzada")
        print(f"{'='*80}")
        print(f"[BUSCAR_PRESUPUESTOS] Filtros recibidos:")
        print(f"  - numero: {p_numero}")
        print(f"  - estado: {p_estado}")
        print(f"  - id_proyecto: {p_id_proyecto}")
        print(f"  - id_obra: {p_id_obra}")
        print(f"  - fecha_desde: {p_fecha_desde}")
        print(f"  - fecha_hasta: {p_fecha_hasta}")
        print(f"  - nombre_creador: {p_nombre_creador}")
        print(f"  - monto_desde: {p_monto_desde}")
        print(f"  - monto_hasta: {p_monto_hasta}")
        print(f"{'='*80}")
        
        try:
            # Query directa (SP roto, usar query directa)
            where_clauses = ["pr.estado != 'ELIMINADO'"]
            params = []
            
            if p_numero:
                where_clauses.append("pr.numero_presupuesto LIKE %s")
                params.append(f"%{p_numero}%")
            if p_estado:
                where_clauses.append("pr.estado = %s")
                params.append(p_estado)
            if p_id_proyecto:
                where_clauses.append("o.id_proyecto = %s")
                params.append(p_id_proyecto)
            if p_id_obra:
                where_clauses.append("pr.id_obra = %s")
                params.append(p_id_obra)
            if p_fecha_desde:
                where_clauses.append("pr.fecha_creacion >= %s")
                params.append(p_fecha_desde)
            if p_fecha_hasta:
                where_clauses.append("pr.fecha_creacion <= %s")
                params.append(p_fecha_hasta + ' 23:59:59')
            if p_monto_desde:
                where_clauses.append("pr.monto >= %s")
                params.append(p_monto_desde)
            if p_monto_hasta:
                where_clauses.append("pr.monto <= %s")
                params.append(p_monto_hasta)
            
            where_sql = " AND ".join(where_clauses)
            
            cursor.execute(f'''
                SELECT 
                    pr.id_presupuesto,
                    pr.numero_presupuesto,
                    pr.estado,
                    pr.monto,
                    pr.fecha_actualizacion,
                    pr.fecha_creacion,
                    o.nombre as nombre_obra,
                    p.nombre as nombre_proyecto,
                    CONCAT(
                        COALESCE(per.nombres, ''),
                        ' ',
                        COALESCE(per.apellido_paterno, ''),
                        ' ',
                        COALESCE(per.apellido_materno, '')
                    ) as creado_por
                FROM TblPresupuesto pr
                LEFT JOIN TblObra o ON pr.id_obra = o.id_obra
                LEFT JOIN TblProyecto p ON o.id_proyecto = p.id_proyecto
                LEFT JOIN TblUsuario u ON pr.num_documento = u.num_documento
                LEFT JOIN TblPersona per ON u.num_documento = per.num_documento
                WHERE {where_sql}
                ORDER BY pr.fecha_creacion DESC
            ''', tuple(params))
            
            presupuestos = cursor.fetchall()
            
            print(f"[BUSCAR_PRESUPUESTOS] [OK] {len(presupuestos)} registros encontrados")
            
            cursor.close()
            connection.close()
            
            return jsonify({'success': True, 'data': presupuestos}), 200
        
        except Error as db_error:
            print(f"[BUSCAR_PRESUPUESTOS] [ERROR] Error en SP: {db_error}")
            print(f"[BUSCAR_PRESUPUESTOS] [ERROR] Cdigo: {getattr(db_error, 'errno', 'N/A')}")
            print(f"[BUSCAR_PRESUPUESTOS] [ERROR] SQL State: {getattr(db_error, 'sqlstate', 'N/A')}")
            print(f"{'='*80}\n")
            
            cursor.close()
            if connection:
                connection.close()
            
            return jsonify({'success': False, 'error': f'Error en base de datos: {str(db_error)}'}), 500
    
    except Exception as e:
        print(f"[BUSCAR_PRESUPUESTOS] [ERROR] Excepcin general: {e}")
        import traceback
        print(traceback.format_exc())
        print(f"{'='*80}\n")
        
        if connection:
            connection.close()
        
        return jsonify({'success': False, 'error': f'Error del servidor: {str(e)}'}), 500

@main_bp.route('/api/presupuestos/obtener/<int:id_presupuesto>', methods=['GET'])
@login_required
def obtener_presupuesto(id_presupuesto):
    """Obtener datos de un presupuesto especfico sin id_material"""
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'error': 'Error de conexin'}), 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        print(f"\n{'='*80}")
        print(f"[OBTENER_PRESUPUESTO] Iniciando para ID: {id_presupuesto}")
        print(f"{'='*80}")
        
        cursor.execute("""
            SELECT 
                pr.id_presupuesto,
                pr.numero_presupuesto,
                pr.id_obra,
                pr.num_documento,
                o.codigo_obra,
                o.nombre as nombre_obra,
                p.codigo_proyecto,
                p.nombre as nombre_proyecto,
                per.nombres as usuario_nombre,
                per.apellido_paterno as usuario_apellido,
                pr.monto,
                pr.estado,
                pr.observaciones,
                pr.fecha_creacion,
                pr.fecha_actualizacion,
                COUNT(DISTINCT pd.id_detalle) as total_items,
                GROUP_CONCAT(DISTINCT m.nombre SEPARATOR ', ') as materiales_utilizados
            FROM TblPresupuesto pr
            INNER JOIN TblObra o ON pr.id_obra = o.id_obra
            INNER JOIN TblProyecto p ON o.id_proyecto = p.id_proyecto
            INNER JOIN TblUsuario u ON pr.num_documento = u.num_documento
            INNER JOIN TblPersona per ON u.num_documento = per.num_documento
            LEFT JOIN TblPresupuestoDetalle pd ON pr.id_presupuesto = pd.id_presupuesto
            LEFT JOIN TblMateriales m ON pd.id_material = m.id_material
            WHERE pr.id_presupuesto = %s
            GROUP BY pr.id_presupuesto, pr.numero_presupuesto, pr.id_obra, pr.num_documento,
                     o.codigo_obra, o.nombre, p.codigo_proyecto, p.nombre,
                     per.nombres, per.apellido_paterno, pr.monto, pr.estado,
                     pr.observaciones, pr.fecha_creacion, pr.fecha_actualizacion
        """, (id_presupuesto,))
        
        presupuesto = cursor.fetchone()
        
        cursor.close()
        connection.close()
        
        if not presupuesto:
            print(f"[OBTENER_PRESUPUESTO] [WARN] Presupuesto no encontrado")
            return jsonify({'success': False, 'error': 'Presupuesto no encontrado'}), 404
        
        print(f"[OBTENER_PRESUPUESTO] [OK] Presupuesto encontrado")
        print(f"{'='*80}\n")
        
        return jsonify({'success': True, 'data': presupuesto}), 200
    
    except Error as e:
        print(f"[OBTENER_PRESUPUESTO] [ERROR] Error SQL: {e}")
        print(f"{'='*80}\n")
        return jsonify({'success': False, 'error': str(e)}), 500

@main_bp.route('/api/presupuestos/visualizar/<int:id_presupuesto>', methods=['GET'])
@login_required
def visualizar_presupuesto(id_presupuesto):
    """Obtener detalles completos del presupuesto usando SP"""
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'error': 'Error de conexin'}), 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        print(f"\n{'='*80}")
        print(f"[VISUALIZAR_PRESUPUESTO] Llamando SP para ID: {id_presupuesto}")
        print(f"{'='*80}")
        
        # Llamar al SP directamente
        print(f"[VISUALIZAR_PRESUPUESTO] [DEBUG] Ejecutando SP directamente...")
        cursor.execute(f'CALL sp_obtener_presupuesto_detalle_completo({id_presupuesto})', multi=True)
        
        # El SP retorna 3 conjuntos de resultados: PRESUPUESTO, DETALLES, RESUMEN
        presupuesto_data = None
        detalles_data = None
        resumen_data = None
        
        # PARTE 1: Obtener presupuesto
        try:
            resultado1 = cursor.fetchall()
            if resultado1:
                presupuesto_data = resultado1[0]
                print(f"[VISUALIZAR_PRESUPUESTO] [DEBUG] PARTE 1 (Presupuesto): {presupuesto_data}")
            else:
                print(f"[VISUALIZAR_PRESUPUESTO] [DEBUG] PARTE 1 vaca")
        except Exception as e:
            print(f"[VISUALIZAR_PRESUPUESTO] [DEBUG] Error PARTE 1: {e}")
        
        # PARTE 2: Obtener detalles
        try:
            if cursor.nextset():
                detalles_data = cursor.fetchall()
                print(f"[VISUALIZAR_PRESUPUESTO] [DEBUG] PARTE 2 (Detalles): {len(detalles_data or [])} registros")
            else:
                print(f"[VISUALIZAR_PRESUPUESTO] [DEBUG] No hay PARTE 2")
        except Exception as e:
            print(f"[VISUALIZAR_PRESUPUESTO] [DEBUG] Error PARTE 2: {e}")
        
        # PARTE 3: Obtener resumen
        try:
            if cursor.nextset():
                resumen_result = cursor.fetchall()
                if resumen_result:
                    resumen_data = resumen_result[0]
                    print(f"[VISUALIZAR_PRESUPUESTO] [DEBUG] PARTE 3 (Resumen): {resumen_data}")
                else:
                    print(f"[VISUALIZAR_PRESUPUESTO] [DEBUG] PARTE 3 vaca")
            else:
                print(f"[VISUALIZAR_PRESUPUESTO] [DEBUG] No hay PARTE 3")
        except Exception as e:
            print(f"[VISUALIZAR_PRESUPUESTO] [DEBUG] Error PARTE 3: {e}")
        
        cursor.close()
        connection.close()
        
        if not presupuesto_data:
            print(f"[VISUALIZAR_PRESUPUESTO] [WARN] Presupuesto no encontrado (PARTE 1 vaca)")
            print(f"[VISUALIZAR_PRESUPUESTO] [INFO] Detalles encontrados: {len(detalles_data or [])} items")
            print(f"[VISUALIZAR_PRESUPUESTO] [INFO] Resumen: {resumen_data}")
            return jsonify({'success': False, 'error': 'Presupuesto no encontrado'}), 404
        
        # Estructurar respuesta
        respuesta = {
            'presupuesto': presupuesto_data,
            'detalles': detalles_data or [],
            'resumen': resumen_data or {}
        }
        
        print(f"[VISUALIZAR_PRESUPUESTO] [OK] Presupuesto obtenido con {len(detalles_data or [])} items")
        print(f"{'='*80}\n")
        
        return jsonify({'success': True, 'data': respuesta}), 200
    
    except Error as e:
        print(f"[VISUALIZAR_PRESUPUESTO] [ERROR] Error SQL: {e}")
        print(f"{'='*80}\n")
        return jsonify({'success': False, 'error': str(e)}), 500
    except Exception as e:
        print(f"[VISUALIZAR_PRESUPUESTO] [ERROR] Error general: {e}")
        print(f"{'='*80}\n")
        return jsonify({'success': False, 'error': str(e)}), 500

@main_bp.route('/api/presupuestos/crear', methods=['POST'])
@login_required
def crear_presupuesto():
    """Crear un nuevo presupuesto con mltiples materiales y servicios usando SP"""
    connection = None
    print(f"\n{'='*80}")
    print(f"[CREAR_PRESUPUESTO] Iniciando...")
    print(f"{'='*80}")
    
    try:
        # Obtener num_documento del usuario autenticado
        num_documento = session.get('user_documento')
        print(f"[CREAR_PRESUPUESTO] Usuario autenticado: {num_documento}")
        
        if not num_documento:
            print(f"[CREAR_PRESUPUESTO] [ERROR] Usuario no autenticado o num_documento no disponible")
            return jsonify({'success': False, 'error': 'Usuario no autenticado'}), 401
        
        datos = request.get_json()
        if not datos:
            print(f"[CREAR_PRESUPUESTO] [ERROR] No se recibi JSON vlido")
            return jsonify({'success': False, 'error': 'Datos invlidos'}), 400
        
        print(f"[CREAR_PRESUPUESTO] [OK] Datos recibidos del frontend")
        
        # Mostrar TODOS los datos recibidos
        print(f"[CREAR_PRESUPUESTO] ")
        print(f"[CREAR_PRESUPUESTO] DATOS DEL FRONTEND:")
        print(f"[CREAR_PRESUPUESTO] ")
        for key, value in datos.items():
            if key in ['materiales', 'servicios']:
                print(f"[CREAR_PRESUPUESTO]   {key}: {len(value)} items")
                if len(value) > 0 and isinstance(value, list):
                    print(f"[CREAR_PRESUPUESTO]      Primer item: {value[0]}")
            else:
                print(f"[CREAR_PRESUPUESTO]   {key}: {value}")
        
        # Validar datos obligatorios
        if not datos.get('id_empresa') or not datos.get('id_obra'):
            print(f"[CREAR_PRESUPUESTO] [ERROR] Falta id_empresa o id_obra")
            return jsonify({'success': False, 'error': 'Empresa y Obra son requeridas'}), 400
        
        # Validar que al menos haya un material o servicio
        materiales = datos.get('materiales', [])
        servicios = datos.get('servicios', [])
        
        if not materiales and not servicios:
            print(f"[CREAR_PRESUPUESTO] [ERROR] Sin materiales ni servicios")
            return jsonify({'success': False, 'error': 'Debe agregar al menos un material o servicio'}), 400
        
        print(f"[CREAR_PRESUPUESTO] [OK] Validaciones bsicas pasadas")
        
        connection = get_db_connection()
        if not connection:
            print(f"[CREAR_PRESUPUESTO] [ERROR] No se pudo conectar a BD")
            return jsonify({'success': False, 'error': 'Error de conexin a base de datos'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # Preparar JSON para materiales y servicios
        import json
        materiales_json = json.dumps(materiales)
        servicios_json = json.dumps(servicios)
        
        print(f"[CREAR_PRESUPUESTO] ")
        print(f"[CREAR_PRESUPUESTO] DATOS A ENVIAR AL SP:")
        print(f"[CREAR_PRESUPUESTO] ")
        print(f"[CREAR_PRESUPUESTO] Parmetro 1: id_empresa = {datos.get('id_empresa')}")
        print(f"[CREAR_PRESUPUESTO] Parmetro 2: id_obra = {datos.get('id_obra')}")
        print(f"[CREAR_PRESUPUESTO] Parmetro 3: comentarios = '{datos.get('comentarios', '')}'")
        print(f"[CREAR_PRESUPUESTO] Parmetro 4: materiales_json = {len(materiales_json)} bytes, {len(materiales)} items")
        print(f"[CREAR_PRESUPUESTO] Parmetro 5: servicios_json = {len(servicios_json)} bytes, {len(servicios)} items")
        
        if len(materiales) > 0:
            print(f"[CREAR_PRESUPUESTO] Ejemplo material: {json.dumps(materiales[0], indent=2)}")
        if len(servicios) > 0:
            print(f"[CREAR_PRESUPUESTO] Ejemplo servicio: {json.dumps(servicios[0], indent=2)}")
        
        print(f"[CREAR_PRESUPUESTO] ")
        print(f"[CREAR_PRESUPUESTO] Ejecutando SP: sp_CrearPresupuestoCompleto")
        print(f"[CREAR_PRESUPUESTO] ")
        
        try:
            # Obtener valores de desglose (con valores por defecto si no existen)
            gastos_generales = float(datos.get('gastos_generales', 0))
            utilidad = float(datos.get('utilidad', 0))
            supervision_obra = float(datos.get('supervision_obra', 0))
            
            print(f"[CREAR_PRESUPUESTO] Desglose:")
            print(f"  - Gastos Generales: {gastos_generales}")
            print(f"  - Utilidad: {utilidad}")
            print(f"  - Supervisin Obra: {supervision_obra}")
            
            # Llamar SP con los parmetros (incluyendo num_documento del usuario autenticado)
            cursor.execute("""
                CALL sp_CrearPresupuestoCompleto(
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    @p_id_presupuesto_created
                )
            """, (
                int(datos.get('id_empresa')),
                int(datos.get('id_obra')),
                num_documento,
                datos.get('comentarios', ''),
                gastos_generales,
                utilidad,
                supervision_obra,
                materiales_json,
                servicios_json
            ))
            
            print(f"[CREAR_PRESUPUESTO] [OK] SP ejecutado correctamente")
            
            # Obtener el ID del presupuesto creado
            cursor.execute("SELECT @p_id_presupuesto_created as id_presupuesto")
            result = cursor.fetchone()
            
            if not result:
                print(f"[CREAR_PRESUPUESTO] [ERROR] No se obtuvo ID de presupuesto")
                raise Exception("No se pudo obtener el ID del presupuesto creado")
            
            id_presupuesto = result.get('id_presupuesto')
            
            if not id_presupuesto:
                print(f"[CREAR_PRESUPUESTO] [ERROR] ID presupuesto es NULL: {result}")
                raise Exception("El SP no retorn un ID vlido")
            
            connection.commit()
            
            print(f"[CREAR_PRESUPUESTO] [OK] [OK] Presupuesto creado con ID: {id_presupuesto}")
            print(f"[CREAR_PRESUPUESTO]   - Materiales insertados: {len(materiales)}")
            print(f"[CREAR_PRESUPUESTO]   - Servicios insertados: {len(servicios)}")
            print(f"[CREAR_PRESUPUESTO]   - Usuario: {num_documento}")
            print(f"{'='*80}\n")
            
            cursor.close()
            connection.close()
            
            return jsonify({
                'success': True,
                'message': 'Presupuesto creado correctamente',
                'id_presupuesto': id_presupuesto,
                'data': {
                    'id_presupuesto_creado': id_presupuesto
                }
            }), 201
        
        except Error as db_error:
            print(f"[CREAR_PRESUPUESTO] [ERROR] Error en SP: {str(db_error)}")
            print(f"[CREAR_PRESUPUESTO] Cdigo: {getattr(db_error, 'errno', 'N/A')}")
            print(f"[CREAR_PRESUPUESTO] SQL State: {getattr(db_error, 'sqlstate', 'N/A')}")
            print(f"[CREAR_PRESUPUESTO] Message: {getattr(db_error, 'msg', 'N/A')}")
            
            if connection:
                connection.rollback()
                cursor.close()
                connection.close()
            
            print(f"{'='*80}\n")
            return jsonify({'success': False, 'error': f'Error en base de datos: {str(db_error)}'}), 500
    
    except Exception as e:
        print(f"[CREAR_PRESUPUESTO] [X] [ERROR] Excepcin general: {str(e)}")
        import traceback
        print(f"[CREAR_PRESUPUESTO] Traceback:")
        print(traceback.format_exc())
        print(f"{'='*80}\n")
        
        if connection:
            try:
                connection.rollback()
                connection.close()
            except:
                pass
        
        return jsonify({'success': False, 'error': f'Error del servidor: {str(e)}'}), 500

@main_bp.route('/api/presupuestos/eliminar/<int:id_presupuesto>', methods=['DELETE'])
@login_required
def eliminar_presupuesto(id_presupuesto):
    """Eliminar un presupuesto (soft delete)"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Error de conexin'}), 500
        
        cursor = connection.cursor()
        
        try:
            print(f"\n{'='*80}")
            print(f"[ELIMINAR_PRESUPUESTO] Iniciando para ID: {id_presupuesto}")
            print(f"{'='*80}")
            
            # Usar SP que elimina presupuesto Y sus registros de aprobacin
            cursor.callproc('sp_MarcarPresupuestoEliminado', [id_presupuesto])
            
            # Obtener resultado del SP
            for result in cursor.stored_results():
                sp_result = result.fetchone()
                if sp_result and sp_result[0] == 'ERROR':
                    raise Error(sp_result[1])
            
            print(f"[ELIMINAR_PRESUPUESTO] [OK] Presupuesto y registros de aprobacin eliminados")
            
            connection.commit()
            cursor.close()
            connection.close()
            
            print(f"{'='*80}\n")
            
            return jsonify({
                'success': True,
                'message': 'Presupuesto eliminado exitosamente'
            }), 200
        
        except Error as e:
            connection.rollback()
            cursor.close()
            connection.close()
            print(f"[ELIMINAR_PRESUPUESTO] [ERROR] Error SQL: {e}")
            print(f"{'='*80}\n")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    except Exception as e:
        print(f"[ELIMINAR_PRESUPUESTO] [ERROR] Error general: {e}")
        print(f"{'='*80}\n")
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================================================
# ENDPOINTS PARA CARGAR COMBO DATA (OBRAS, USUARIOS, MATERIALES)
# ============================================================================

@main_bp.route('/api/presupuestos/combo/obras', methods=['GET'])
@login_required
def combo_obras():
    """Obtener lista de obras para dropdown"""
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'error': 'Error de conexin'}), 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                o.id_obra,
                o.codigo_obra,
                o.nombre,
                p.nombre as nombre_proyecto
            FROM TblObra o
            INNER JOIN TblProyecto p ON o.id_proyecto = p.id_proyecto
            ORDER BY o.codigo_obra
        """)
        
        obras = cursor.fetchall()
        cursor.close()
        connection.close()
        
        return jsonify({'success': True, 'data': obras}), 200
    
    except Error as e:
        print(f"[COMBO_OBRAS] [ERROR] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@main_bp.route('/api/presupuestos/combo/usuarios', methods=['GET'])
@login_required
def combo_usuarios():
    """Obtener lista de usuarios para dropdown"""
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'error': 'Error de conexin'}), 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                u.num_documento,
                p.nombres,
                p.apellido_paterno,
                p.apellido_materno,
                u.usuario
            FROM TblUsuario u
            INNER JOIN TblPersona p ON u.num_documento = p.num_documento
            WHERE u.estado != 'INACTIVO'
            ORDER BY p.nombres
        """)
        
        usuarios = cursor.fetchall()
        cursor.close()
        connection.close()
        
        return jsonify({'success': True, 'data': usuarios}), 200
    
    except Error as e:
        print(f"[COMBO_USUARIOS] [ERROR] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@main_bp.route('/api/presupuestos/combo/categorias', methods=['GET'])
@login_required
def combo_categorias():
    """Obtener lista de categoras de materiales desde SP"""
    print(f"\n[COMBO_CATEGORIAS] Iniciando...")
    
    connection = None
    try:
        connection = get_db_connection()
        if not connection:
            print(f"[COMBO_CATEGORIAS] [ERROR] No se pudo conectar a BD")
            return jsonify({'success': False, 'error': 'Error de conexin a base de datos'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        print(f"[COMBO_CATEGORIAS] Llamando SP: sp_ObtenerCategoriasMaterial()")
        # Llamar al SP con callproc
        try:
            cursor.callproc('sp_ObtenerCategoriasMaterial')
            
            # Obtener resultados del SP
            categorias = []
            for result in cursor.stored_results():
                categorias = result.fetchall()
                break
            
            print(f"[COMBO_CATEGORIAS] [OK] {len(categorias)} categorías obtenidas")
            
            cursor.close()
            connection.close()
            
            return jsonify({'success': True, 'data': categorias}), 200
        
        except Error as db_error:
            print(f"[COMBO_CATEGORIAS] [ERROR] Error en SP: {db_error}")
            print(f"[COMBO_CATEGORIAS] [ERROR] Cdigo: {getattr(db_error, 'errno', 'N/A')}, SQL State: {getattr(db_error, 'sqlstate', 'N/A')}")
            cursor.close()
            if connection:
                connection.close()
            return jsonify({'success': False, 'error': f'Error en base de datos: {str(db_error)}'}), 500
    
    except Exception as e:
        print(f"[COMBO_CATEGORIAS] [ERROR] Excepcin general: {e}")
        import traceback
        print(f"[COMBO_CATEGORIAS] [TRACEBACK]:\n{traceback.format_exc()}")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': f'Error del servidor: {str(e)}'}), 500

@main_bp.route('/api/presupuestos/combo/materiales', methods=['GET'])
@login_required
def combo_materiales():
    """Obtener lista de materiales para dropdown - con filtros dinámicos"""
    print(f"\n[COMBO_MATERIALES] Iniciando...")
    
    # Obtener parámetros de búsqueda
    termino_busqueda = request.args.get('termino', '').strip()
    id_categoria = request.args.get('categoria', '0')
    
    print(f"[COMBO_MATERIALES] Parámetros recibidos:")
    print(f"  - termino_busqueda: '{termino_busqueda}' (tipo: {type(termino_busqueda).__name__}, len: {len(termino_busqueda)})")
    print(f"  - id_categoria: '{id_categoria}' (tipo: {type(id_categoria).__name__})")
    
    # Si el término está vacío, retornar array vacío (no llamar al SP)
    if not termino_busqueda or len(termino_busqueda) == 0:
        print(f"[COMBO_MATERIALES] [INFO] Término vacío, retornando array vacío")
        return jsonify({'success': True, 'data': []}), 200
    
    connection = None
    try:
        connection = get_db_connection()
        if not connection:
            print(f"[COMBO_MATERIALES] [ERROR] No se pudo conectar a BD")
            return jsonify({'success': False, 'error': 'Error de conexión a base de datos'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # Convertir id_categoria a int
        try:
            id_categoria_int = int(id_categoria)
        except (ValueError, TypeError) as e:
            print(f"[COMBO_MATERIALES] [ERROR] id_categoria inválido: {e}")
            cursor.close()
            connection.close()
            return jsonify({'success': False, 'error': f'Parámetro categoria inválido: {id_categoria}'}), 400
        
        print(f"[COMBO_MATERIALES] Llamando SP: sp_BuscarMateriales(termino='{termino_busqueda}', categoria={id_categoria_int})")
        
        # Llamar al SP con parámetros
        try:
            cursor.callproc('sp_BuscarMateriales', (termino_busqueda, id_categoria_int))
            
            # Obtener resultados
            materiales = []
            for result in cursor.stored_results():
                materiales = result.fetchall()
                break
            
            print(f"[COMBO_MATERIALES] [OK] {len(materiales)} materiales encontrados")
            
            cursor.close()
            connection.close()
            
            return jsonify({'success': True, 'data': materiales}), 200
        
        except Error as db_error:
            print(f"[COMBO_MATERIALES] [ERROR] Error en ejecución de SP: {db_error}")
            print(f"[COMBO_MATERIALES] [ERROR] Código de error: {getattr(db_error, 'errno', 'N/A')}")
            print(f"[COMBO_MATERIALES] [ERROR] SQL State: {getattr(db_error, 'sqlstate', 'N/A')}")
            
            cursor.close()
            if connection:
                connection.close()
            
            return jsonify({'success': False, 'error': f'Error en base de datos: {str(db_error)}'}), 500
    
    except Exception as e:
        print(f"[COMBO_MATERIALES] [ERROR] Excepción general: {e}")
        import traceback
        print(f"[COMBO_MATERIALES] [TRACEBACK]:\n{traceback.format_exc()}")
        
        if connection:
            connection.close()
        
        return jsonify({'success': False, 'error': f'Error del servidor: {str(e)}'}), 500

# ============================================================================
# ENDPOINTS PARA TblPresupuestoDetalle (ITEMS DE MATERIALES)
# ============================================================================

@main_bp.route('/api/presupuesto-detalle/crear', methods=['POST'])
@login_required
def crear_presupuesto_detalle():
    """Crear un item de material para un presupuesto"""
    try:
        data = request.get_json()
        
        # Validar campos - id_material es OPCIONAL
        campos_requeridos = ['id_presupuesto', 'cantidad', 'precio_unitario']
        for campo in campos_requeridos:
            if not data.get(campo):
                return jsonify({'success': False, 'error': f'Campo requerido: {campo}'}), 400
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Error de conexin'}), 500
        
        try:
            cursor = connection.cursor()
            
            print(f"[CREAR_PRESUPUESTO_DETALLE] Creando item para presupuesto {data['id_presupuesto']}")
            
            # Insertar item - id_material puede ser NULL
            cursor.execute("""
                INSERT INTO TblPresupuestoDetalle (
                    id_presupuesto,
                    id_material,
                    cantidad,
                    precio_unitario,
                    observaciones,
                    fecha_creacion
                ) VALUES (%s, %s, %s, %s, %s, NOW())
            """, (
                data['id_presupuesto'],
                data.get('id_material'),  # Puede ser None/null
                data['cantidad'],
                data['precio_unitario'],
                data.get('observaciones', '')
            ))
            
            detalle_id = cursor.lastrowid
            connection.commit()
            cursor.close()
            connection.close()
            
            print(f"[CREAR_PRESUPUESTO_DETALLE] [OK] Item creado con ID: {detalle_id}")
            
            return jsonify({
                'success': True,
                'message': 'Item creado exitosamente',
                'id': detalle_id
            }), 201
        
        except Error as e:
            connection.rollback()
            cursor.close()
            connection.close()
            print(f"[CREAR_PRESUPUESTO_DETALLE] [ERROR] {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    except Exception as e:
        print(f"[CREAR_PRESUPUESTO_DETALLE] [ERROR] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@main_bp.route('/api/presupuesto-detalle/<int:id_presupuesto>', methods=['GET'])
@login_required
def obtener_presupuesto_detalle(id_presupuesto):
    """Obtener items de un presupuesto"""
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'error': 'Error de conexin'}), 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                pd.id_detalle,
                pd.id_presupuesto,
                pd.id_material,
                pd.cantidad,
                pd.precio_unitario,
                pd.subtotal,
                pd.observaciones,
                m.codigo_material,
                m.nombre as nombre_material
            FROM TblPresupuestoDetalle pd
            INNER JOIN TblMateriales m ON pd.id_material = m.id_material
            WHERE pd.id_presupuesto = %s
            ORDER BY pd.fecha_creacion
        """, (id_presupuesto,))
        
        detalles = cursor.fetchall()
        cursor.close()
        connection.close()
        
        return jsonify({'success': True, 'data': detalles}), 200
    
    except Error as e:
        print(f"[OBTENER_PRESUPUESTO_DETALLE] [ERROR] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@main_bp.route('/api/presupuesto-detalle/eliminar/<int:id_detalle>', methods=['DELETE'])
@login_required
def eliminar_presupuesto_detalle(id_detalle):
    """Eliminar un item de presupuesto"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Error de conexin'}), 500
        
        cursor = connection.cursor()
        
        try:
            print(f"[ELIMINAR_PRESUPUESTO_DETALLE] Eliminando item: {id_detalle}")
            
            cursor.execute("DELETE FROM TblPresupuestoDetalle WHERE id_detalle = %s", (id_detalle,))
            
            connection.commit()
            cursor.close()
            connection.close()
            
            print(f"[ELIMINAR_PRESUPUESTO_DETALLE] [OK] Item eliminado")
            
            return jsonify({
                'success': True,
                'message': 'Item eliminado exitosamente'
            }), 200
        
        except Error as e:
            connection.rollback()
            cursor.close()
            connection.close()
            print(f"[ELIMINAR_PRESUPUESTO_DETALLE] [ERROR] {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    except Exception as e:
        print(f"[ELIMINAR_PRESUPUESTO_DETALLE] [ERROR] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================================================
# NUEVOS ENDPOINTS PARA CARGAR EMPRESAS, PROYECTOS Y OBRAS
# ============================================================================

@main_bp.route('/api/presupuestos/empresas/listar', methods=['GET'])
@login_required
def presupuestos_listar_empresas():
    """Obtener lista de empresas desde TblEmpresa para presupuestos"""
    print(f"\n[PRESUPUESTOS_LISTAR_EMPRESAS] Iniciando...")
    
    connection = get_db_connection()
    if not connection:
        print(f"[PRESUPUESTOS_LISTAR_EMPRESAS] [ERROR] No se pudo conectar a BD")
        return jsonify({'success': False, 'error': 'Error de conexin a base de datos'}), 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        print(f"[PRESUPUESTOS_LISTAR_EMPRESAS] Consultando TblEmpresa directamente")
        cursor.execute("""
            SELECT id_empresa, nombre 
            FROM TblEmpresa 
            WHERE activa = 1 
            ORDER BY nombre ASC
        """)
        
        empresas = cursor.fetchall()
        print(f"[PRESUPUESTOS_LISTAR_EMPRESAS] [OK] {len(empresas)} empresas obtenidas")
        
        cursor.close()
        connection.close()
        
        return jsonify({'success': True, 'data': empresas}), 200
    
    except Error as e:
        print(f"[PRESUPUESTOS_LISTAR_EMPRESAS] [ERROR] Error MySQL: {str(e)}")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': f'Error MySQL: {str(e)}'}), 500
    except Exception as e:
        print(f"[PRESUPUESTOS_LISTAR_EMPRESAS] [ERROR] Error general: {str(e)}")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': f'Error: {str(e)}'}), 500


@main_bp.route('/api/presupuestos/proyectos/listar', methods=['GET'])
@login_required
def presupuestos_listar_proyectos():
    """Obtener lista de proyectos desde TblProyecto para presupuestos"""
    print(f"\n[PRESUPUESTOS_LISTAR_PROYECTOS] Iniciando...")
    
    connection = get_db_connection()
    if not connection:
        print(f"[PRESUPUESTOS_LISTAR_PROYECTOS] [ERROR] No se pudo conectar a BD")
        return jsonify({'success': False, 'error': 'Error de conexin a base de datos'}), 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        print(f"[PRESUPUESTOS_LISTAR_PROYECTOS] Consultando TblProyecto directamente")
        cursor.execute("""
            SELECT id_proyecto, nombre 
            FROM TblProyecto 
            ORDER BY nombre ASC
        """)
        
        proyectos = cursor.fetchall()
        print(f"[PRESUPUESTOS_LISTAR_PROYECTOS] [OK] {len(proyectos)} proyectos obtenidos")
        
        cursor.close()
        connection.close()
        
        return jsonify({'success': True, 'data': proyectos}), 200
    
    except Error as e:
        print(f"[PRESUPUESTOS_LISTAR_PROYECTOS] [ERROR] Error MySQL: {str(e)}")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': f'Error MySQL: {str(e)}'}), 500
    except Exception as e:
        print(f"[PRESUPUESTOS_LISTAR_PROYECTOS] [ERROR] Error general: {str(e)}")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': f'Error: {str(e)}'}), 500


@main_bp.route('/api/presupuestos/obras/listar', methods=['GET'])
@login_required
def presupuestos_listar_obras():
    """Obtener lista de obras filtradas por proyecto para presupuestos"""
    print(f"\n[PRESUPUESTOS_LISTAR_OBRAS] Iniciando...")
    
    id_proyecto = request.args.get('id_proyecto')
    print(f"[PRESUPUESTOS_LISTAR_OBRAS] id_proyecto: {id_proyecto}")
    
    if not id_proyecto:
        print(f"[PRESUPUESTOS_LISTAR_OBRAS] [ERROR] id_proyecto es requerido")
        return jsonify({'success': False, 'error': 'id_proyecto es requerido'}), 400
    
    connection = get_db_connection()
    if not connection:
        print(f"[PRESUPUESTOS_LISTAR_OBRAS] [ERROR] No se pudo conectar a BD")
        return jsonify({'success': False, 'error': 'Error de conexin a base de datos'}), 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        print(f"[PRESUPUESTOS_LISTAR_OBRAS] Consultando TblObra directamente para proyecto {id_proyecto}")
        cursor.execute("""
            SELECT id_obra, nombre 
            FROM TblObra 
            WHERE id_proyecto = %s 
            ORDER BY nombre ASC
        """, (int(id_proyecto),))
        
        obras = cursor.fetchall()
        print(f"[PRESUPUESTOS_LISTAR_OBRAS] [OK] {len(obras)} obras obtenidas")
        
        cursor.close()
        connection.close()
        
        return jsonify({'success': True, 'data': obras}), 200
    
    except Error as e:
        print(f"[PRESUPUESTOS_LISTAR_OBRAS] [ERROR] Error MySQL: {str(e)}")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': f'Error MySQL: {str(e)}'}), 500
    except Exception as e:
        print(f"[PRESUPUESTOS_LISTAR_OBRAS] [ERROR] Error general: {str(e)}")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': f'Error: {str(e)}'}), 500

# ============================================================================
# ENDPOINTS PARA EDITAR PRESUPUESTOS
# ============================================================================

@main_bp.route('/api/presupuestos/obtener-para-editar/<int:id_presupuesto>', methods=['GET'])
@login_required
def obtener_presupuesto_para_editar(id_presupuesto):
    """Obtener encabezado y detalles de presupuesto para editar - FORMATO COMPATIBLE CON FRONTEND"""
    print(f"\n{'='*80}")
    print(f"[OBTENER_PRESUPUESTO_PARA_EDITAR] ID: {id_presupuesto}")
    print(f"{'='*80}")
    
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'error': 'Error de conexin'}), 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        # 1. Obtener informacin bsica del presupuesto
        cursor.execute("""
            SELECT 
                p.id_presupuesto,
                p.numero_presupuesto,
                p.id_empresa,
                e.nombre as nombre_empresa,
                p.id_obra,
                p.num_documento,
                COALESCE(p.monto_total, 0) as monto_total,
                COALESCE(p.monto_aprobado, 0) as monto_aprobado,
                COALESCE(p.gastos_generales, 0) as gastos_generales,
                COALESCE(p.utilidad, 0) as utilidad,
                COALESCE(p.igv, 0) as igv,
                COALESCE(p.supervision_obra, 0) as supervision_obra,
                p.estado,
                p.observaciones,
                o.id_proyecto,
                o.nombre as nombre_obra,
                pr.nombre as nombre_proyecto
            FROM TblPresupuesto p
            LEFT JOIN TblEmpresa e ON p.id_empresa = e.id_empresa
            LEFT JOIN TblObra o ON p.id_obra = o.id_obra
            LEFT JOIN TblProyecto pr ON o.id_proyecto = pr.id_proyecto
            WHERE p.id_presupuesto = %s
        """, (id_presupuesto,))
        
        presupuesto = cursor.fetchone()
        if not presupuesto:
            print(f"[OBTENER_PRESUPUESTO_PARA_EDITAR] Presupuesto no encontrado")
            return jsonify({'success': False, 'error': 'Presupuesto no encontrado'}), 404
        
        print(f"[OBTENER_PRESUPUESTO_PARA_EDITAR] [OK] Presupuesto encontrado")
        
        # 2. Obtener detalles (materiales y servicios)
        cursor.execute("""
            SELECT 
                pd.id_detalle,
                pd.id_presupuesto,
                pd.id_material,
                pd.tipo_item,
                pd.descripcion,
                pd.cantidad,
                pd.cantidad_original,
                pd.cantidad_consumida,
                COALESCE(pd.cantidad_saldo, 0) as cantidad_saldo,
                pd.precio_unitario,
                pd.subtotal,
                COALESCE(m.nombre, pd.descripcion) as nombre_item,
                m.codigo_material,
                COALESCE(mc.nombre, 'General') as categoria,
                COALESCE(um.nombre, 'und') as unidad_medida,
                COALESCE(um.nombre, 'und') as unidad_nombre,
                COALESCE(um.abreviatura, 'und') as unidad_abreviatura
            FROM TblPresupuestoDetalle pd
            LEFT JOIN TblMateriales m ON pd.id_material = m.id_material
            LEFT JOIN TblCategoriaMaterial mc ON m.id_categoria = mc.id_categoria
            LEFT JOIN TblUnidadMedida um ON m.id_unidad = um.id_unidad
            WHERE pd.id_presupuesto = %s
            ORDER BY pd.tipo_item ASC, pd.id_detalle ASC
        """, (id_presupuesto,))
        
        detalles_raw = cursor.fetchall()
        print(f"[OBTENER_PRESUPUESTO_PARA_EDITAR] [OK] {len(detalles_raw) if detalles_raw else 0} detalles obtenidos")
        
        # 3. Procesar detalles en arrays separados (formato que espera el frontend)
        materiales = []
        servicios = []
        
        if detalles_raw:
            for d in detalles_raw:
                if d.get('tipo_item') == 'SERVICIO':
                    servicios.append({
                        'id_temporal': d['id_detalle'],
                        'id_detalle': d['id_detalle'],
                        'descripcion': d.get('descripcion', ''),
                        'cantidad': float(d.get('cantidad', 0)),
                        'precio_unitario': float(d.get('precio_unitario', 0)),
                        'subtotal': float(d.get('subtotal', 0))
                    })
                else:  # MATERIAL
                    materiales.append({
                        'id_temporal': d['id_detalle'],
                        'id_detalle': d['id_detalle'],
                        'id_material': d.get('id_material'),
                        'nombre': d.get('nombre_item', 'Sin nombre'),
                        'codigo': d.get('codigo_material', ''),
                        'categoria': d.get('categoria', 'General'),
                        'unidad': d.get('unidad_medida', 'und'),
                        'cantidad': float(d.get('cantidad', 0)),
                        'precio_unitario': float(d.get('precio_unitario', 0)),
                        'subtotal': float(d.get('subtotal', 0))
                    })
        
        print(f"[OBTENER_PRESUPUESTO_PARA_EDITAR] [OK] {len(materiales)} materiales, {len(servicios)} servicios")
        
        cursor.close()
        connection.close()
        
        # Retornar en formato que espera el frontend
        respuesta = {
            'presupuesto': presupuesto,
            'data': {
                'detalles': materiales + servicios,
                'materiales': materiales,
                'servicios': servicios
            }
        }
        
        print(f"[OBTENER_PRESUPUESTO_PARA_EDITAR] [OK] Enviando respuesta")
        print(f"{'='*80}\n")
        
        return jsonify({'success': True, 'data': respuesta}), 200
    
    except Exception as e:
        print(f"[OBTENER_PRESUPUESTO_PARA_EDITAR] [ERROR] {str(e)}")
        import traceback
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


@main_bp.route('/api/presupuestos/obtener-detalles/<int:id_presupuesto>', methods=['GET'])
@login_required
def obtener_presupuesto_detalles(id_presupuesto):
    """Obtener detalles completos del presupuesto"""
    print(f"\n{'='*80}")
    print(f"[OBTENER_PRESUPUESTO_DETALLES] Iniciando... id_presupuesto={id_presupuesto}")
    print(f"{'='*80}")
    
    connection = get_db_connection()
    if not connection:
        print(f"[OBTENER_PRESUPUESTO_DETALLES] [ERROR] No se pudo conectar a BD")
        return jsonify({'success': False, 'error': 'Error de conexin'}), 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        # 1. Obtener informacin del presupuesto
        print(f"[OBTENER_PRESUPUESTO_DETALLES] Obteniendo informacin del presupuesto...")
        cursor.execute("""
            SELECT 
                pr.id_presupuesto,
                pr.numero_presupuesto,
                pr.id_obra,
                pr.num_documento,
                o.codigo_obra,
                o.nombre as nombre_obra,
                p.codigo_proyecto,
                p.nombre as nombre_proyecto,
                per.nombres as usuario_nombre,
                per.apellido_paterno as usuario_apellido,
                pr.monto,
                pr.estado,
                pr.observaciones,
                pr.fecha_creacion,
                pr.fecha_actualizacion
            FROM TblPresupuesto pr
            INNER JOIN TblObra o ON pr.id_obra = o.id_obra
            INNER JOIN TblProyecto p ON o.id_proyecto = p.id_proyecto
            INNER JOIN TblUsuario u ON pr.num_documento = u.num_documento
            INNER JOIN TblPersona per ON u.num_documento = per.num_documento
            WHERE pr.id_presupuesto = %s
        """, (id_presupuesto,))
        
        presupuesto = cursor.fetchone()
        if not presupuesto:
            print(f"[OBTENER_PRESUPUESTO_DETALLES] [ERROR] Presupuesto no encontrado")
            return jsonify({'success': False, 'error': 'Presupuesto no encontrado'}), 404
        
        print(f"[OBTENER_PRESUPUESTO_DETALLES] [OK] Presupuesto obtenido")
        
        # 2. Obtener detalles (materiales y servicios)
        print(f"[OBTENER_PRESUPUESTO_DETALLES] Obteniendo detalles...")
        cursor.execute("""
            SELECT 
                pd.id_detalle,
                pd.id_presupuesto,
                pd.id_material,
                pd.cantidad,
                pd.cantidad_original,
                pd.cantidad_consumida,
                COALESCE(pd.cantidad_saldo, 0) as cantidad_saldo,
                pd.precio_unitario,
                pd.subtotal,
                pd.observaciones,
                pd.descripcion,
                COALESCE(pd.tipo_item, 'MATERIAL') as tipo_item,
                COALESCE(m.codigo_material, '') as codigo_material,
                COALESCE(m.nombre, pd.descripcion, 'Sin nombre') as nombre_item,
                COALESCE(um.nombre, 'und') as unidad_medida,
                COALESCE(um.nombre, 'und') as unidad_nombre,
                COALESCE(um.abreviatura, 'und') as unidad_abreviatura,
                COALESCE(mc.nombre, 'General') as categoria
            FROM TblPresupuestoDetalle pd
            LEFT JOIN TblMateriales m ON pd.id_material = m.id_material
            LEFT JOIN TblUnidadMedida um ON m.id_unidad = um.id_unidad
            LEFT JOIN TblCategoriaMaterial mc ON m.id_categoria = mc.id_categoria
            WHERE pd.id_presupuesto = %s
            ORDER BY COALESCE(pd.tipo_item, 'MATERIAL'), pd.id_detalle
        """, (id_presupuesto,))
        
        detalles = cursor.fetchall()
        print(f"[OBTENER_PRESUPUESTO_DETALLES] [OK] {len(detalles) if detalles else 0} detalles obtenidos")
        
        # 3. Calcular resumen
        materiales = [d for d in detalles if d.get('tipo_item') == 'MATERIAL'] if detalles else []
        servicios = [d for d in detalles if d.get('tipo_item') == 'SERVICIO'] if detalles else []
        
        total_materiales = sum(float(m.get('subtotal', 0) or 0) for m in materiales) if materiales else 0
        total_servicios = sum(float(s.get('subtotal', 0) or 0) for s in servicios) if servicios else 0
        total_general = total_materiales + total_servicios
        
        resumen = {
            'cantidad_items': len(detalles) if detalles else 0,
            'cantidad_materiales': len(materiales),
            'cantidad_servicios': len(servicios),
            'total_materiales': total_materiales,
            'total_servicios': total_servicios,
            'monto_total_calculado': total_general
        }
        
        print(f"[OBTENER_PRESUPUESTO_DETALLES] [OK] Resumen calculado: {resumen}")
        print(f"[OBTENER_PRESUPUESTO_DETALLES] Materiales: {len(materiales)}, Servicios: {len(servicios)}")
        print(f"{'='*80}\n")
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'data': {
                'presupuesto': presupuesto,
                'detalles': detalles,
                'resumen': resumen
            }
        }), 200
    
    except Error as e:
        print(f"[OBTENER_PRESUPUESTO_DETALLES] [ERROR] Error MySQL: {str(e)}")
        print(f"{'='*80}\n")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': f'Error MySQL: {str(e)}'}), 500
    except Exception as e:
        print(f"[OBTENER_PRESUPUESTO_DETALLES] [ERROR] Error general: {str(e)}")
        print(f"{'='*80}\n")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': f'Error: {str(e)}'}), 500


# ============================================================================
# ENDPOINT PARA ACTUALIZAR PRESUPUESTOS (EDICIN)
# ============================================================================

@main_bp.route('/api/presupuestos/actualizar/<int:id_presupuesto>', methods=['PUT'])
@login_required
def actualizar_presupuesto_editar(id_presupuesto):
    """Actualizar presupuesto existente usando SP"""
    import sys
    
    print(f"\n{'='*80}", flush=True)
    print(f"[ACTUALIZAR_PRESUPUESTO] Iniciando para ID: {id_presupuesto}", flush=True)
    print(f"{'='*80}", flush=True)
    sys.stdout.flush()
    
    try:
        datos = request.get_json()
        
        print(f"\n[ACTUALIZAR_PRESUPUESTO] DATOS RECIBIDOS DEL FRONTEND:", flush=True)
        print(f"{''*80}", flush=True)
        print(f"  id_empresa: {datos.get('id_empresa')}", flush=True)
        print(f"  id_proyecto: {datos.get('id_proyecto')}", flush=True)
        print(f"  id_obra: {datos.get('id_obra')}", flush=True)
        print(f"  comentarios: {datos.get('comentarios', '(vaco)')}", flush=True)
        print(f"  cantidad_materiales: {len(datos.get('materiales', []))}", flush=True)
        print(f"  cantidad_servicios: {len(datos.get('servicios', []))}", flush=True)
        print(f"{''*80}", flush=True)
        
        if datos.get('materiales'):
            print(f"\n[ACTUALIZAR_PRESUPUESTO] DETALLES DE MATERIALES:", flush=True)
            for idx, m in enumerate(datos.get('materiales', []), 1):
                print(f"  [{idx}] id_material: {m.get('id_material')} | nombre: {m.get('nombre')} | cantidad: {m.get('cantidad')} | precio: {m.get('precio_unitario')} | subtotal: {m.get('subtotal')}", flush=True)
        else:
            print(f"\n[ACTUALIZAR_PRESUPUESTO] SIN MATERIALES", flush=True)
        
        if datos.get('servicios'):
            print(f"\n[ACTUALIZAR_PRESUPUESTO] DETALLES DE SERVICIOS:", flush=True)
            for idx, s in enumerate(datos.get('servicios', []), 1):
                print(f"  [{idx}] descripcion: {s.get('descripcion')} | cantidad: {s.get('cantidad')} | precio: {s.get('precio_unitario')} | subtotal: {s.get('subtotal')}", flush=True)
        else:
            print(f"\n[ACTUALIZAR_PRESUPUESTO] SIN SERVICIOS", flush=True)
        
        sys.stdout.flush()
        
        connection = get_db_connection()
        if not connection:
            print(f"\n[ACTUALIZAR_PRESUPUESTO] [ERROR] No se pudo conectar a BD", flush=True)
            print(f"{'='*80}\n", flush=True)
            return jsonify({'success': False, 'error': 'Error de conexin'}), 500
        
        cursor = connection.cursor()
        
        # Preparar JSON para materiales y servicios
        import json
        
        # Limpiar descripciones de servicios ANTES de enviar al SP
        servicios_limpios = []
        for servicio in datos.get('servicios', []):
            servicio_limpio = servicio.copy()
            # Limpiar escapes JSON de la descripcin
            if 'descripcion' in servicio_limpio and servicio_limpio['descripcion']:
                desc = servicio_limpio['descripcion']
                # Remover escapes innecesarios
                desc = desc.replace('\\"', '"').replace('\\\\', '\\')
                servicio_limpio['descripcion'] = desc
            servicios_limpios.append(servicio_limpio)
        
        materiales_json = json.dumps(datos.get('materiales', []), ensure_ascii=False)
        servicios_json = json.dumps(servicios_limpios, ensure_ascii=False)
        
        print(f"\n[ACTUALIZAR_PRESUPUESTO] JSON A ENVIAR AL SP:", flush=True)
        print(f"{''*80}", flush=True)
        print(f"  materiales_json ({len(materiales_json)} caracteres):", flush=True)
        print(f"  CONTENIDO COMPLETO:", flush=True)
        print(f"  {materiales_json}", flush=True)
        print(f"\n  servicios_json ({len(servicios_json)} caracteres):", flush=True)
        print(f"  CONTENIDO COMPLETO:", flush=True)
        print(f"  {servicios_json}", flush=True)
        print(f"{''*80}", flush=True)
        
        print(f"\n[ACTUALIZAR_PRESUPUESTO] PARMETROS QUE SE ENVIARN AL SP:", flush=True)
        print(f"  1. id_presupuesto = {id_presupuesto} (tipo: {type(id_presupuesto).__name__})", flush=True)
        print(f"  2. id_empresa = {datos.get('id_empresa')} (tipo: {type(datos.get('id_empresa')).__name__})", flush=True)
        print(f"  3. id_obra = {datos.get('id_obra')} (tipo: {type(datos.get('id_obra')).__name__})", flush=True)
        print(f"  4. comentarios = '{datos.get('comentarios', '')}' (largo: {len(datos.get('comentarios', ''))} caracteres)", flush=True)
        print(f"  5. materiales_json = {len(materiales_json)} caracteres", flush=True)
        print(f"  6. servicios_json = {len(servicios_json)} caracteres", flush=True)
        
        print(f"\n[ACTUALIZAR_PRESUPUESTO] Ejecutando: CALL sp_ActualizarPresupuestoCompleto(...)", flush=True)
        sys.stdout.flush()
        sys.stdout.flush()
        
        # Obtener num_documento del usuario autenticado
        num_documento = session.get('user_documento')
        
        # Llamar SP
        print(f"\n[ACTUALIZAR_PRESUPUESTO] Ejecutando consulta SQL...", flush=True)
        cursor.execute("""
            CALL sp_ActualizarPresupuestoCompleto(
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s
            )
        """, (
            id_presupuesto,
            datos.get('id_empresa'),
            datos.get('id_obra'),
            num_documento,
            datos.get('comentarios', ''),
            float(datos.get('gastos_generales', 0)),
            float(datos.get('utilidad', 0)),
            float(datos.get('supervision_obra', 0)),
            materiales_json,
            servicios_json
        ))
        
        print(f"[ACTUALIZAR_PRESUPUESTO] Consulta ejecutada, haciendo COMMIT...", flush=True)
        connection.commit()
        print(f"[ACTUALIZAR_PRESUPUESTO] COMMIT realizado", flush=True)
        
        # PASO ADICIONAL: REINICIAR FLUJO DE APROBACIN USANDO SP
        print(f"\n[ACTUALIZAR_PRESUPUESTO] REINICIANDO FLUJO DE APROBACIN (usando SP)...", flush=True)
        
        try:
            # Llamar SP para reiniciar flujo
            cursor.callproc('sp_ReiniciarFlujoAprobacion', [1, id_presupuesto])
            
            # Obtener resultado del SP
            resultado_sp = cursor.fetchone()
            if resultado_sp:
                resultado = resultado_sp[0]
                mensaje = resultado_sp[1] if len(resultado_sp) > 1 else ""
                pasos = resultado_sp[2] if len(resultado_sp) > 2 else 0
                
                if resultado == 'OK':
                    print(f"[ACTUALIZAR_PRESUPUESTO]   [OK] {mensaje}", flush=True)
                else:
                    print(f"[ACTUALIZAR_PRESUPUESTO]   [!] {mensaje}", flush=True)
            
            connection.commit()
            print(f"[ACTUALIZAR_PRESUPUESTO]   [OK] Flujo reiniciado", flush=True)
        
        except Exception as e:
            print(f"[ACTUALIZAR_PRESUPUESTO]   [!] Error reiniciando flujo: {e}", flush=True)
            connection.rollback()
        
        print(f"\n[ACTUALIZAR_PRESUPUESTO] [[OK] OK] Presupuesto actualizado y flujo reiniciado", flush=True)
        print(f"[ACTUALIZAR_PRESUPUESTO] RESUMEN:", flush=True)
        print(f"  ID Presupuesto: {id_presupuesto}", flush=True)
        print(f"  Empresa: {datos.get('id_empresa')}", flush=True)
        print(f"  Obra: {datos.get('id_obra')}", flush=True)
        print(f"  Materiales insertados: {len(datos.get('materiales', []))}", flush=True)
        print(f"  Servicios insertados: {len(datos.get('servicios', []))}", flush=True)
        print(f"{'='*80}\n", flush=True)
        sys.stdout.flush()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': 'Presupuesto actualizado correctamente',
            'data': {
                'id_presupuesto': id_presupuesto,
                'id_empresa': datos.get('id_empresa'),
                'id_obra': datos.get('id_obra'),
                'materiales_count': len(datos.get('materiales', [])),
                'servicios_count': len(datos.get('servicios', []))
            }
        }), 200
    
    except Exception as e:
        print(f"\n[ACTUALIZAR_PRESUPUESTO] [[X] ERROR] {str(e)}", flush=True)
        import traceback
        print(traceback.format_exc(), flush=True)
        print(f"{'='*80}\n", flush=True)
        sys.stdout.flush()
        if connection:
            try:
                connection.rollback()
                connection.close()
            except:
                pass
        return jsonify({'success': False, 'error': f'Error: {str(e)}'}), 500


# ============================================================================
# ENDPOINT PARA BSQUEDA DE PRESUPUESTOS
# ============================================================================

@main_bp.route('/api/presupuestos/buscar-por-numero', methods=['GET'])
@login_required
def buscar_presupuesto_por_numero():
    """Buscar presupuesto por nmero - Retorna cantidad_saldo en lugar de monto_disponible"""
    numero = request.args.get('numero', '').strip()
    
    print(f"\n{'='*80}")
    print(f"[BUSCAR_PRESUPUESTO] Buscando por nmero: {numero}")
    print(f"{'='*80}")
    
    if not numero:
        return jsonify({'success': False, 'error': 'Nmero de presupuesto requerido'}), 400
    
    connection = get_db_connection()
    if not connection:
        print(f"[BUSCAR_PRESUPUESTO] [ERROR] No se pudo conectar a BD")
        return jsonify({'success': False, 'error': 'Error de conexin'}), 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        # TASK 6: Return cantidad_saldo instead of monto_disponible
        cursor.execute("""
            SELECT 
                pr.id_presupuesto,
                pr.numero_presupuesto,
                pr.id_obra,
                pr.num_documento,
                o.codigo_obra,
                o.nombre as nombre_obra,
                p.codigo_proyecto,
                p.nombre as nombre_proyecto,
                pr.monto,
                COALESCE(pr.cantidad_consumida, 0) as cantidad_consumida,
                COALESCE(pr.cantidad_saldo, 0) as cantidad_saldo,
                pr.estado,
                pr.observaciones,
                pr.fecha_creacion
            FROM TblPresupuesto pr
            INNER JOIN TblObra o ON pr.id_obra = o.id_obra
            INNER JOIN TblProyecto p ON o.id_proyecto = p.id_proyecto
            WHERE pr.numero_presupuesto = %s
              AND pr.estado = 'APROBADO'
        """, (numero,))
        
        presupuesto = cursor.fetchone()
        
        cursor.close()
        connection.close()
        
        if presupuesto:
            print(f"[BUSCAR_PRESUPUESTO] [OK] Presupuesto encontrado")
            print(f"[BUSCAR_PRESUPUESTO]   Monto: {presupuesto['monto']}")
            print(f"[BUSCAR_PRESUPUESTO]   Cantidad Consumida: {presupuesto['cantidad_consumida']}")
            print(f"[BUSCAR_PRESUPUESTO]   Cantidad Saldo: {presupuesto['cantidad_saldo']}")
            print(f"{'='*80}\n")
            return jsonify({'success': True, 'data': presupuesto}), 200
        else:
            print(f"[BUSCAR_PRESUPUESTO] [X] Presupuesto no encontrado")
            print(f"{'='*80}\n")
            return jsonify({'success': False, 'error': 'Presupuesto no encontrado'}), 404
    
    except Error as e:
        print(f"[BUSCAR_PRESUPUESTO] [ERROR] {str(e)}")
        print(f"{'='*80}\n")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': f'Error MySQL: {str(e)}'}), 500
    except Exception as e:
        print(f"[BUSCAR_PRESUPUESTO] [ERROR] {str(e)}")
        print(f"{'='*80}\n")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': f'Error: {str(e)}'}), 500


# ============================================================================

@main_bp.route('/api/presupuestos/aprobar/<int:id_presupuesto>', methods=['POST'])
@login_required
def aprobar_presupuesto(id_presupuesto):
    """Aprobar presupuesto - Cambiar estado a APROBADO"""
    num_documento = session.get('user_documento')
    
    print(f"\n{'='*80}")
    print(f"[APROBAR_PRESUPUESTO] Iniciando... id_presupuesto={id_presupuesto}")
    print(f"[APROBAR_PRESUPUESTO] Usuario: {num_documento}")
    print(f"{'='*80}")
    
    connection = get_db_connection()
    if not connection:
        print(f"[APROBAR_PRESUPUESTO] [ERROR] No se pudo conectar a BD")
        return jsonify({'success': False, 'error': 'Error de conexin'}), 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        print(f"[APROBAR_PRESUPUESTO] Llamando SP: sp_AprobarPresupuesto_Progresivo({id_presupuesto}, {num_documento}, 1)")
        
        # Usar SP progresivo con 3 parmetros: id_presupuesto, usuario, tipo_documento
        cursor.callproc('sp_AprobarPresupuesto_Progresivo', [id_presupuesto, num_documento, 1])
        
        # Consumir el resultado
        resultado = None
        for result in cursor.stored_results():
            rows = result.fetchall()
            if rows:
                resultado = rows[0]
                break
        
        print(f"[APROBAR_PRESUPUESTO] [OK] Presupuesto aprobado")
        print(f"[APROBAR_PRESUPUESTO] Resultado: {resultado}")
        
        # CRITICAL: Commit the transaction
        print(f"[APROBAR_PRESUPUESTO] Haciendo COMMIT...")
        connection.commit()
        print(f"[APROBAR_PRESUPUESTO] COMMIT realizado [OK]")
        print(f"{'='*80}\n")
        
        cursor.close()
        connection.close()
        
        if resultado and resultado.get('resultado') == 'OK':
            return jsonify({
                'success': True,
                'message': 'Presupuesto aprobado correctamente'
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': resultado.get('mensaje', 'Error desconocido') if resultado else 'Error ejecutando SP'
            }), 400
    
    except Error as e:
        print(f"[APROBAR_PRESUPUESTO] [ERROR] Error MySQL: {str(e)}")
        print(f"{'='*80}\n")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': f'Error MySQL: {str(e)}'}), 500
    except Exception as e:
        print(f"[APROBAR_PRESUPUESTO] [ERROR] Error general: {str(e)}")
        print(f"{'='*80}\n")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': f'Error: {str(e)}'}), 500


# ============================================================================
# ENDPOINT PARA RECHAZAR PRESUPUESTO
# ============================================================================

@main_bp.route('/api/presupuestos/rechazar/<int:id_presupuesto>', methods=['POST'])
@login_required
def rechazar_presupuesto(id_presupuesto):
    """Rechazar presupuesto - Cambiar estado a RECHAZADO y registrar motivo"""
    num_documento = session.get('user_documento')
    
    print(f"\n{'='*80}")
    print(f"[RECHAZAR_PRESUPUESTO] Iniciando... id_presupuesto={id_presupuesto}")
    print(f"[RECHAZAR_PRESUPUESTO] Usuario: {num_documento}")
    print(f"{'='*80}")
    
    # Obtener motivo del request body - OPCIONAL
    datos = request.get_json() or {}
    motivo = datos.get('motivo', '') or None
    
    print(f"[RECHAZAR_PRESUPUESTO] Motivo: {motivo or '(sin motivo)'}")
    
    connection = get_db_connection()
    if not connection:
        print(f"[RECHAZAR_PRESUPUESTO] [ERROR] No se pudo conectar a BD")
        return jsonify({'success': False, 'error': 'Error de conexin'}), 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        print(f"[RECHAZAR_PRESUPUESTO] Llamando SP: sp_RechazarPresupuesto_Progresivo({id_presupuesto}, {num_documento}, '{motivo}', 1)")
        
        # Llamar SP progresivo con 4 parmetros: id, usuario que rechaza, motivo, tipo_documento
        cursor.callproc('sp_RechazarPresupuesto_Progresivo', [id_presupuesto, num_documento, motivo, 1])
        
        # Consumir el resultado
        resultado = None
        for result in cursor.stored_results():
            rows = result.fetchall()
            if rows:
                resultado = rows[0]
                break
        
        print(f"[RECHAZAR_PRESUPUESTO] [OK] Presupuesto rechazado")
        print(f"[RECHAZAR_PRESUPUESTO] Resultado: {resultado}")
        
        # CRITICAL: Commit the transaction
        print(f"[RECHAZAR_PRESUPUESTO] Haciendo COMMIT...")
        connection.commit()
        print(f"[RECHAZAR_PRESUPUESTO] COMMIT realizado [OK]")
        print(f"{'='*80}\n")
        
        cursor.close()
        connection.close()
        
        if resultado and resultado.get('resultado') == 'OK':
            return jsonify({
                'success': True,
                'message': 'Presupuesto rechazado correctamente'
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': resultado.get('mensaje', 'Error desconocido') if resultado else 'Error ejecutando SP'
            }), 400
    
    except Error as e:
        print(f"[RECHAZAR_PRESUPUESTO] [ERROR] Error MySQL: {str(e)}")
        print(f"{'='*80}\n")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': f'Error MySQL: {str(e)}'}), 500
    except Exception as e:
        print(f"[RECHAZAR_PRESUPUESTO] [ERROR] Error general: {str(e)}")
        print(f"{'='*80}\n")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': f'Error: {str(e)}'}), 500
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': f'Error MySQL: {str(e)}'}), 500
    except Exception as e:
        print(f"[RECHAZAR_PRESUPUESTO] [ERROR] Error general: {str(e)}")
        print(f"{'='*80}\n")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': f'Error: {str(e)}'}), 500

# ============================================================================
# ENDPOINT PARA OBTENER FLUJO DE APROBACIN PROGRESIVO
# ============================================================================

@main_bp.route('/api/presupuestos/flujo/<int:id_presupuesto>', methods=['GET'])
@login_required
def obtener_flujo_aprobacion(id_presupuesto):
    """Obtener flujo de aprobacin y historial de un presupuesto"""
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'error': 'Error de conexin'}), 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        print(f"\n{'='*80}")
        print(f"[FLUJO_APROBACION] Obteniendo flujo para presupuesto: {id_presupuesto}")
        print(f"{'='*80}")
        
        # PASO 1: Obtener informacin del presupuesto
        cursor.execute("""
            SELECT 
                id_presupuesto,
                numero_presupuesto,
                estado,
                monto,
                fecha_creacion
            FROM TblPresupuesto
            WHERE id_presupuesto = %s
        """, (id_presupuesto,))
        
        presupuesto = cursor.fetchone()
        
        if not presupuesto:
            print(f"[FLUJO_APROBACION] [WARN] Presupuesto no encontrado")
            cursor.close()
            connection.close()
            return jsonify({'success': False, 'error': 'Presupuesto no encontrado'}), 404
        
        print(f"[FLUJO_APROBACION] [OK] Presupuesto encontrado: {presupuesto['numero_presupuesto']}")
        
        # PASO 2: Obtener pasos de aprobacin desde TblFlujoAprobacionCargos
        cursor.execute("""
            SELECT 
                numero_paso,
                nombre_paso,
                descripcion,
                es_final,
                id_cargo
            FROM TblFlujoAprobacionCargos
            WHERE id_tipo_documento = 1
            AND activo = 1
            ORDER BY numero_paso
        """)
        
        pasos = cursor.fetchall()
        print(f"[FLUJO_APROBACION] [OK] {len(pasos)} pasos encontrados")
        
        # PASO 3: Obtener historial de aprobaciones
        cursor.execute("""
            SELECT 
                numero_paso,
                id_cargo_aprobador,
                num_documento_aprobador,
                estado_aprobacion,
                fecha_aprobacion,
                comentario
            FROM TblRegistroAprobacion
            WHERE id_documento_referencia = %s
            AND id_tipo_documento = 1
            ORDER BY numero_paso
        """, (id_presupuesto,))
        
        historial = cursor.fetchall()
        print(f"[FLUJO_APROBACION] [OK] {len(historial)} registros de aprobacin encontrados")
        
        # PASO 4: Obtener nombres de aprobadores del historial
        for registro in historial:
            if registro['num_documento_aprobador']:
                cursor.execute("""
                    SELECT 
                        CONCAT(
                            COALESCE(nombres, ''),
                            ' ',
                            COALESCE(apellido_paterno, ''),
                            ' ',
                            COALESCE(apellido_materno, '')
                        ) as nombre_completo
                    FROM TblPersona
                    WHERE num_documento = %s
                """, (registro['num_documento_aprobador'],))
                
                persona = cursor.fetchone()
                if persona:
                    registro['nombre_aprobador'] = persona['nombre_completo'].strip()
        
        cursor.close()
        connection.close()
        
        # Estruturar respuesta
        respuesta = {
            'presupuesto': presupuesto,
            'flujo': {
                'pasos': pasos,
                'pasos_totales': len(pasos)
            },
            'historial': historial
        }
        
        print(f"[FLUJO_APROBACION] [OK] Flujo obtenido correctamente")
        print(f"{'='*80}\n")
        
        return jsonify({'success': True, 'data': respuesta}), 200
    
    except Error as e:
        print(f"[FLUJO_APROBACION] [ERROR] Error SQL: {e}")
        print(f"{'='*80}\n")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': str(e)}), 500
    except Exception as e:
        print(f"[FLUJO_APROBACION] [ERROR] Error general: {e}")
        print(f"{'='*80}\n")
        if connection:
            connection.close()
        return jsonify({'success': False, 'error': str(e)}), 500


@main_bp.route('/api/presupuestos/obtener-flujo-aprobacion/<int:id_presupuesto>', methods=['GET'])
@login_required
def obtener_flujo_aprobacion_presupuesto(id_presupuesto):
    """
    Obtener el estado de aprobacin por cargo para un presupuesto especfico
    Retorna los crculos del flujo de aprobacin con estado y usuario aprobador
    """
    connection = get_db_connection()
    if not connection:
        return jsonify({'success': False, 'error': 'Error de conexin'}), 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        print(f"\n{'='*80}")
        print(f"[FLUJO_APROBACION_PRESUPUESTO] Obteniendo flujo para presupuesto {id_presupuesto}")
        print(f"{'='*80}")
        
        # Verificar que el presupuesto existe
        cursor.execute("""
            SELECT id_presupuesto, numero_presupuesto, estado
            FROM TblPresupuesto
            WHERE id_presupuesto = %s
        """, (id_presupuesto,))
        
        pres = cursor.fetchone()
        if not pres:
            cursor.close()
            connection.close()
            return jsonify({'success': False, 'error': 'Presupuesto no encontrado'}), 404
        
        # Obtener todos los pasos del flujo de Presupuesto (tipo_documento = 1)
        cursor.execute("""
            SELECT 
                fac.numero_paso,
                fac.nombre_paso,
                fac.id_cargo,
                c.nombre as cargo_nombre,
                a.nombre as area_nombre
            FROM TblFlujoAprobacionCargos fac
            LEFT JOIN TblCargo c ON fac.id_cargo = c.id_cargo
            LEFT JOIN TblArea a ON c.id_area = a.id_area
            WHERE fac.id_tipo_documento = 1
            AND fac.activo = 1
            AND fac.es_requerido = 1
            ORDER BY fac.numero_paso ASC
        """)
        
        pasos_config = cursor.fetchall()
        print(f"   Pasos configurados: {len(pasos_config)}")
        
        # Para cada paso, obtener el estado de aprobacin
        pasos = []
        for paso_config in pasos_config:
            # Buscar registro de aprobacin para este paso
            cursor.execute("""
                SELECT 
                    ra.estado_aprobacion,
                    ra.num_documento_aprobador,
                    ra.fecha_aprobacion,
                    per.nombres,
                    per.apellido_paterno,
                    ra.comentario
                FROM TblRegistroAprobacion ra
                LEFT JOIN TblUsuario u ON ra.num_documento_aprobador = u.num_documento
                LEFT JOIN TblPersona per ON u.num_documento = per.num_documento
                WHERE ra.id_tipo_documento = 1
                AND ra.id_documento_referencia = %s
                AND ra.numero_paso = %s
                LIMIT 1
            """, (id_presupuesto, paso_config['numero_paso']))
            
            aprobacion = cursor.fetchone()
            
            # Determinar estado y usuario
            estado = 'PENDIENTE'
            usuario_aprobador = None
            fecha_aprobacion = None
            comentario = None
            
            if aprobacion:
                estado = aprobacion['estado_aprobacion']
                fecha_aprobacion = aprobacion['fecha_aprobacion']
                comentario = aprobacion.get('comentario')
                
                print(f"   [DEBUG] Paso {paso_config['numero_paso']}: estado={estado}, doc={aprobacion['num_documento_aprobador']}")
                
                if aprobacion['num_documento_aprobador'] and aprobacion['nombres']:
                    usuario_aprobador = f"{aprobacion['nombres']} {aprobacion['apellido_paterno']}".strip()
                    print(f"   [DEBUG]    Usuario: {usuario_aprobador}")
            
            pasos.append({
                'numero_paso': paso_config['numero_paso'],
                'nombre_paso': paso_config['nombre_paso'],
                'id_cargo': paso_config['id_cargo'],
                'cargo_nombre': paso_config['cargo_nombre'],
                'area_nombre': paso_config['area_nombre'],
                'estado': estado,
                'usuario_aprobador': usuario_aprobador,
                'fecha_aprobacion': fecha_aprobacion.isoformat() if fecha_aprobacion else None,
                'comentario': comentario
            })
            
            print(f"   Paso {paso_config['numero_paso']}: {estado} ({usuario_aprobador or 'Sin aprobador'})")
        
        print(f"[FLUJO_APROBACION_PRESUPUESTO] [OK] Flujo obtenido exitosamente")
        print(f"[FLUJO_APROBACION_PRESUPUESTO] Total pasos: {len(pasos)}")
        print(f"{'='*80}\n")
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'presupuesto': pres['numero_presupuesto'],
            'estado': pres['estado'],
            'pasos': pasos
        }), 200
    
    except Error as e:
        print(f"[FLUJO_APROBACION_PRESUPUESTO] [X] ERROR SQL: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# API: CREAR PROYECTO
# ============================================================================
@main_bp.route('/api/proyectos/crear', methods=['POST'])
@login_required
def crear_proyecto():
    """Crear nuevo proyecto usando SP"""
    try:
        data = request.get_json()
        
        if not data.get('nombre'):
            return jsonify({'success': False, 'error': 'El nombre es obligatorio'}), 400
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Error de conexin'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            print(f"\n{'='*80}")
            print(f"[CREAR_PROYECTO] Iniciando creacin de proyecto")
            print(f"[CREAR_PROYECTO] Datos: {data}")
            print(f"{'='*80}")
            
            # Solo enviar nombre y descripcin (campos que existen en TblProyecto)
            cursor.execute("""
                CALL sp_CrearProyecto(
                    %s, %s,
                    @p_id_proyecto, @p_mensaje
                )
            """, (
                data['nombre'],
                data.get('descripcion')
            ))
            
            cursor.execute("SELECT @p_id_proyecto as id_proyecto, @p_mensaje as mensaje")
            resultado = cursor.fetchone()
            
            # Consumir resultados restantes
            while cursor.nextset():
                pass
            
            connection.commit()
            cursor.close()
            connection.close()
            
            if resultado and resultado['id_proyecto'] and resultado['id_proyecto'] > 0:
                print(f"[CREAR_PROYECTO] [OK] Proyecto creado con ID: {resultado['id_proyecto']}")
                print(f"{'='*80}\n")
                return jsonify({
                    'success': True,
                    'message': resultado['mensaje'],
                    'id_proyecto': resultado['id_proyecto']
                }), 201
            else:
                print(f"[CREAR_PROYECTO] [X] Error: {resultado.get('mensaje', 'Error desconocido')}")
                print(f"{'='*80}\n")
                return jsonify({
                    'success': False,
                    'error': resultado.get('mensaje', 'Error al crear proyecto')
                }), 400
        
        except Error as e:
            print(f"[CREAR_PROYECTO] [X] Error SQL: {e}")
            print(f"{'='*80}\n")
            if connection:
                connection.rollback()
            cursor.close()
            connection.close()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    except Exception as e:
        print(f"[CREAR_PROYECTO] [X] Error general: {e}")
        print(f"{'='*80}\n")
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================================================
# API: CREAR OBRA
# ============================================================================
@main_bp.route('/api/obras/crear', methods=['POST'])
@login_required
def crear_obra():
    """Crear nueva obra usando SP"""
    try:
        data = request.get_json()
        
        if not data.get('id_proyecto') or not data.get('nombre'):
            return jsonify({'success': False, 'error': 'Proyecto y nombre son obligatorios'}), 400
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Error de conexin'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            print(f"\n{'='*80}")
            print(f"[CREAR_OBRA] Iniciando creacin de obra")
            print(f"[CREAR_OBRA] Datos: {data}")
            print(f"{'='*80}")
            
            # Solo enviar campos bsicos que existen en TblObra
            # El codigo_obra se genera automticamente en el SP
            cursor.execute("""
                CALL sp_CrearObra(
                    %s, %s, %s,
                    @p_id_obra, @p_mensaje
                )
            """, (
                data['id_proyecto'],
                data['nombre'],
                data.get('descripcion')
            ))
            
            cursor.execute("SELECT @p_id_obra as id_obra, @p_mensaje as mensaje")
            resultado = cursor.fetchone()
            
            # Consumir resultados restantes
            while cursor.nextset():
                pass
            
            connection.commit()
            cursor.close()
            connection.close()
            
            if resultado and resultado['id_obra'] and resultado['id_obra'] > 0:
                print(f"[CREAR_OBRA] [OK] Obra creada con ID: {resultado['id_obra']}")
                print(f"{'='*80}\n")
                return jsonify({
                    'success': True,
                    'message': resultado['mensaje'],
                    'id_obra': resultado['id_obra']
                }), 201
            else:
                print(f"[CREAR_OBRA] [X] Error: {resultado.get('mensaje', 'Error desconocido')}")
                print(f"{'='*80}\n")
                return jsonify({
                    'success': False,
                    'error': resultado.get('mensaje', 'Error al crear obra')
                }), 400
        
        except Error as e:
            print(f"[CREAR_OBRA] [X] Error SQL: {e}")
            print(f"{'='*80}\n")
            if connection:
                connection.rollback()
            cursor.close()
            connection.close()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    except Exception as e:
        print(f"[CREAR_OBRA] [X] Error general: {e}")
        print(f"{'='*80}\n")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# ENDPOINT: IMPORTAR PDF - Extraer materiales y servicios
# ============================================================================
import os
import tempfile
import pdfplumber
import re

# Unidades de tiempo = SERVICIOS
UNIDADES_SERVICIO = [
    # Servicios especiales (HML, ML, GLB)
    'hm', 'ml', 'glb',
    # Tiempo
    'mes', 'meses', 'dia', 'dias', 'da', 'das',
    'semana', 'semanas', 'hr', 'hora', 'horas', 'h',
    'min', 'minuto', 'minutos', 'ao', 'aos',
    'quincena', 'bimestre', 'trimestre', 'semestre'
]

# Unidades que SIEMPRE son materiales (no dependen de la clasificacion)
UNIDADES_MATERIAL = ['und', 'vol', 'bulto', 'm2', 'm3', 'kg', 'lt', 'gl', 'unid', 'par', 'jgo']

# Palabras clave que indican fila de resumen/totales (NO son items)
FILAS_RESUMEN = [
    'costo directo', 'costo total', 'gastos generales', 'utilidad',
    'sub total', 'subtotal', 'igv', 'presupuesto total',
    'total de ejecucion', 'total general', 'suma total',
    'gravamen', 'impuesto'
]


def clasificar_unidad(unidad_texto):
    """Clasifica una unidad como MATERIAL o SERVICIO"""
    if not unidad_texto:
        return 'MATERIAL'
    u = unidad_texto.strip().lower()
    for serv in UNIDADES_SERVICIO:
        if u == serv:
            return 'SERVICIO'
    return 'MATERIAL'


def es_fila_item(codigo):
    """Verifica si el cdigo tiene formato de item (01.01.01)"""
    return bool(re.match(r'^\d{2}\.\d{2}\.\d{2}', codigo))


def extraer_items_del_pdf(filepath):
    """
    Extrae items del PDF usando texto plano (ms confiable que tablas para este formato).
    Formato esperado: 01.01.01 DESCRIPCION unidad cantidad precio subtotal
    
    Versin mejorada: maneja descripciones en mltiples lneas y extrae porcentajes.
    """
    items = []
    porcentajes = {
        'gastos_generales': 0,
        'utilidad': 0,
        'igv': 18  # Default IGV
    }
    lineas_procesadas = 0
    lineas_descartadas = 0
    
    print(f"[EXTRAER_PDF] Iniciando extraccin de: {filepath}")
    
    with pdfplumber.open(filepath) as pdf:
        print(f"[EXTRAER_PDF] PDF tiene {len(pdf.pages)} pginas")
        
        for num_pagina, pagina in enumerate(pdf.pages):
            texto = pagina.extract_text()
            if not texto:
                print(f"[EXTRAER_PDF] Pgina {num_pagina + 1}: sin texto")
                continue
            
            lineas = texto.split('\n')
            print(f"[EXTRAER_PDF] Pgina {num_pagina + 1}: {len(lineas)} lneas")
            
            # Extraer porcentajes del PDF
            for linea in lineas:
                linea_lower = linea.lower().strip()
                
                # Buscar GASTOS GENERALES (X%)
                match_gg = re.search(r'gastos generales\s*\(\s*(\d+(?:\.\d+)?)\s*%\s*\)', linea_lower)
                if match_gg:
                    porcentajes['gastos_generales'] = float(match_gg.group(1))
                    print(f"[EXTRAER_PDF] Porcentaje Gastos Generales: {porcentajes['gastos_generales']}%")
                
                # Buscar UTILIDAD (X%)
                match_util = re.search(r'utilidad\s*\(\s*(\d+(?:\.\d+)?)\s*%\s*\)', linea_lower)
                if match_util:
                    porcentajes['utilidad'] = float(match_util.group(1))
                    print(f"[EXTRAER_PDF] Porcentaje Utilidad: {porcentajes['utilidad']}%")
                
                # Buscar IGV (X%)
                match_igv = re.search(r'igv\s*\(\s*(\d+(?:\.\d+)?)\s*%\s*\)', linea_lower)
                if match_igv:
                    porcentajes['igv'] = float(match_igv.group(1))
                    print(f"[EXTRAER_PDF] Porcentaje IGV: {porcentajes['igv']}%")
            
            # Primera pasada: identificar qu lneas son items y cules son descripcin
            lineas_items = []  # Guarda (indice_linea, codigo, resto)
            lineas_texto_previo = []  # Lneas de texto puro (sin cdigo)
            
            for idx, linea in enumerate(lineas):
                linea_stripped = linea.strip()
                
                # Verificar si empieza con cdigo XX.XX.XX
                match = re.match(r'^(\d{2}\.\d{2}\.\d{2})\s+(.*)', linea_stripped)
                
                if match:
                    codigo = match.group(1)
                    resto = match.group(2).strip()
                    lineas_items.append((idx, codigo, resto))
                else:
                    # Es lnea de texto (posible descripcin)
                    # Solo si tiene contenido significativo (no es ttulo/subttulo de grupo)
                    if linea_stripped and not re.match(r'^\d{2}\s', linea_stripped):
                        lineas_texto_previo.append((idx, linea_stripped))
            
            # Segunda pasada: procesar cada item
            for idx_item, (num_linea, codigo, resto) in enumerate(lineas_items):
                lineas_procesadas += 1
                
                # Intentar extraer nmeros del resto de la lnea
                match_numeros = re.search(
                    r'(.+?)\s+'                            # descripcin
                    r'([a-zA-Z]+[\d]*)\s+'           # unidad
                    r'([\d,.]+)\s+'                         # cantidad
                    r'([\d,.]+)\s+'                         # precio
                    r'([\d,.]+)\s*$',                       # subtotal
                    resto
                )
                
                if match_numeros:
                    # Caso normal: descripcin + unidad + nmeros en la misma lnea
                    descripcion = match_numeros.group(1).strip()
                    unidad = match_numeros.group(2).strip()
                    cantidad = float(match_numeros.group(3).replace(',', ''))
                    precio = float(match_numeros.group(4).replace(',', ''))
                    subtotal = float(match_numeros.group(5).replace(',', ''))
                else:
                    # Caso especial: la descripcin est en lneas anteriores
                    # Buscar las lneas de texto justo antes de esta lnea de item
                    descripcion_parts = []
                    
                    # Buscar lneas de texto antes de esta lnea del item
                    for idx_texto, texto_linea in reversed(lineas_texto_previo):
                        if idx_texto < num_linea:
                            # Verificar si es lnea de continuacin (no es cdigo ni ttulo)
                            if not re.match(r'^\d{2}\.\d{2}', texto_linea):
                                descripcion_parts.insert(0, texto_linea)
                            else:
                                break
                    
                    # Si encontramos descripcin en lneas anteriores, intentar extraer nmeros
                    if descripcion_parts:
                        descripcion_completa = ' '.join(descripcion_parts) + ' ' + resto
                        
                        match_numeros_extendido = re.search(
                            r'(.+?)\s+'
                            r'([a-zA-Z]+[\d]*)\s+'
                            r'([\d,.]+)\s+'
                            r'([\d,.]+)\s+'
                            r'([\d,.]+)\s*$',
                            descripcion_completa
                        )
                        
                        if match_numeros_extendido:
                            descripcion = match_numeros_extendido.group(1).strip()
                            unidad = match_numeros_extendido.group(2).strip()
                            cantidad = float(match_numeros_extendido.group(3).replace(',', ''))
                            precio = float(match_numeros_extendido.group(4).replace(',', ''))
                            subtotal = float(match_numeros_extendido.group(5).replace(',', ''))
                        else:
                            print(f"[EXTRAER_PDF] Descartado (no se pudieron extraer nmeros): {codigo} - {resto[:60]}")
                            lineas_descartadas += 1
                            continue
                    else:
                        print(f"[EXTRAER_PDF] Descartado (sin descripcin): {codigo} - {resto[:60]}")
                        lineas_descartadas += 1
                        continue
                
                # Filtrar filas de resumen/totales
                desc_lower = descripcion.lower()
                if any(p in desc_lower for p in FILAS_RESUMEN):
                    print(f"[EXTRAER_PDF] Descartado (resumen): {codigo} - {descripcion}")
                    lineas_descartadas += 1
                    continue
                
                # Saltar si la cantidad o precio son 0
                if cantidad <= 0:
                    print(f"[EXTRAER_PDF] Descartado (cantidad=0): {codigo} - {descripcion}")
                    lineas_descartadas += 1
                    continue
                
                # Si no tiene unidad, es titulo/subtitulo (no material ni servicio)
                if not unidad:
                    print(f"[EXTRAER_PDF] Descartado (sin unidad): {codigo} - {descripcion}")
                    lineas_descartadas += 1
                    continue
                
                tipo = clasificar_unidad(unidad)
                
                items.append({
                    'codigo': codigo,
                    'descripcion': descripcion,
                    'cantidad': cantidad,
                    'precio_unitario': precio,
                    'subtotal': subtotal,
                    'unidad': unidad,
                    'tipo': tipo
                })
                
                print(f"[EXTRAER_PDF] [OK] Item: {codigo} - {descripcion[:50]}... | {unidad} | {cantidad} | {precio} | {subtotal}")
    
    print(f"[EXTRAER_PDF] Total: {len(items)} items extrados, {lineas_procesadas} lneas procesadas, {lineas_descartadas} descartadas")
    print(f"[EXTRAER_PDF] Porcentajes extrados: GG={porcentajes['gastos_generales']}%, Util={porcentajes['utilidad']}%, IGV={porcentajes['igv']}%")
    return items, porcentajes


@main_bp.route('/api/presupuestos/importar-pdf', methods=['POST'])
def importar_presupuesto_pdf():
    """
    Recibe un PDF, extrae items, clasifica materiales/servicios.
    Crea automticamente los materiales que no existen en TblMateriales.
    """
    try:
        if 'archivo' not in request.files:
            return jsonify({'success': False, 'error': 'No se proporcion ningn archivo'}), 400
        
        archivo = request.files['archivo']
        
        if not archivo.filename:
            return jsonify({'success': False, 'error': 'Nombre de archivo vaco'}), 400
        
        if not archivo.filename.lower().endswith('.pdf'):
            return jsonify({'success': False, 'error': 'El archivo debe ser un PDF'}), 400
        
        # Guardar temporalmente
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
            archivo.save(tmp.name)
            tmp_path = tmp.name
        
        try:
            # Extraer items del PDF y porcentajes
            items, porcentajes_pdf = extraer_items_del_pdf(tmp_path)
            
            if not items:
                return jsonify({
                    'success': False, 
                    'error': 'No se encontraron items en el PDF. Verifica que el documento contenga lneas con formato: cdigo descripcin unidad cantidad precio subtotal.'
                }), 400
            
            # Conexin a BD para buscar/crear materiales
            connection = get_db_connection()
            if not connection:
                return jsonify({'success': False, 'error': 'Error de conexin a la base de datos'}), 500
            
            try:
                cursor = connection.cursor(dictionary=True)
                
                # Separar materiales y servicios
                materiales = []
                servicios = []
                materiales_nuevos = []
                materiales_existentes = []
                materiales_error = []
                
                print(f"[IMPORTAR_PDF] Procesando {len(items)} items extrados del PDF")
                
                for idx, item in enumerate(items):
                    print(f"[IMPORTAR_PDF] Item {idx + 1}/{len(items)}: {item['descripcion'][:40]}... tipo={item['tipo']}")
                    
                    if item['tipo'] == 'SERVICIO':
                        servicios.append({
                            'descripcion': item['descripcion'],
                            'cantidad': item['cantidad'],
                            'precio_unitario': item['precio_unitario'],
                            'subtotal': item['subtotal'],
                            'unidad': item['unidad']
                        })
                        print(f"[IMPORTAR_PDF]    Servicio agregado")
                    else:
                        nombre_material = item['descripcion']
                        id_material = None
                        
                        # 1. Buscar si el material ya existe (nombre exacto)
                        cursor.execute(
                            "SELECT id_material FROM TblMateriales WHERE nombre = %s AND estado = 'ACTIVO' LIMIT 1",
                            (nombre_material,)
                        )
                        existente = cursor.fetchone()
                        
                        if existente:
                            # Ya existe - usar su ID
                            id_material = existente['id_material']
                            materiales_existentes.append(nombre_material)
                            print(f"[IMPORTAR_PDF]    Material ya existe: ID={id_material}")
                        else:
                            # 2. No existe - buscar id_unidad por abreviatura
                            unidad_pdf = item['unidad']
                            id_unidad = None
                            
                            cursor.execute(
                                """SELECT id_unidad FROM TblUnidadMedida 
                                   WHERE (abreviatura = %s OR nombre = %s OR nombre LIKE %s OR abreviatura LIKE %s)
                                   AND estado = 'ACTIVO' LIMIT 1""",
                                (unidad_pdf, unidad_pdf, f'%{unidad_pdf}%', f'%{unidad_pdf}%')
                            )
                            unidad_result = cursor.fetchone()
                            
                            if unidad_result:
                                id_unidad = unidad_result['id_unidad']
                                print(f"[IMPORTAR_PDF]    Unidad encontrada: {unidad_pdf}  id_unidad={id_unidad}")
                            else:
                                # Si no encuentra unidad, usar unidad por defecto (UND = 1)
                                id_unidad = 1
                                print(f"[IMPORTAR_PDF]    Unidad '{unidad_pdf}' no encontrada, usando UND=1")
                            
                            # 3. Crear material usando SP
                            cursor.execute("SET @p_id = NULL, @p_codigo = NULL, @p_result = 0")
                            cursor.execute(
                                """CALL sp_CrearMaterialConCodigoAuto(
                                    %s, %s, NULL, %s, NULL,
                                    @p_id, @p_codigo, @p_result
                                )""",
                                (nombre_material, item.get('descripcion', nombre_material), id_unidad)
                            )
                            cursor.execute("SELECT @p_id as id, @p_codigo as codigo, @p_result as resultado")
                            sp_result = cursor.fetchone()
                            
                            if sp_result and sp_result['resultado'] == 1:
                                id_material = sp_result['id']
                                materiales_nuevos.append(nombre_material)
                                print(f"[IMPORTAR_PDF]    Material CREADO: {sp_result['codigo']} (ID={id_material})")
                            else:
                                materiales_error.append(nombre_material)
                                print(f"[IMPORTAR_PDF]    ERROR creando material: resultado={sp_result}")
                        
                        materiales.append({
                            'id_material': id_material,
                            'nombre': nombre_material,
                            'cantidad': item['cantidad'],
                            'precio_unitario': item['precio_unitario'],
                            'subtotal': item['subtotal'],
                            'unidad': item['unidad'],
                            'es_nuevo': id_material is not None and nombre_material in materiales_nuevos
                        })
                
                print(f"[IMPORTAR_PDF] [OK] PDF procesado: {len(materiales)} materiales ({len(materiales_nuevos)} nuevos, {len(materiales_existentes)} existentes, {len(materiales_error)} errores), {len(servicios)} servicios")
                
                return jsonify({
                    'success': True,
                    'materiales': materiales,
                    'servicios': servicios,
                    'total_items': len(items),
                    'materiales_nuevos': materiales_nuevos,
                    'materiales_existentes': materiales_existentes,
                    'materiales_error': materiales_error,
                    'porcentajes': porcentajes_pdf,
                    'mensaje': f'Se extrajeron {len(materiales)} materiales ({len(materiales_nuevos)} nuevos, {len(materiales_existentes)} existentes) y {len(servicios)} servicios'
                })
                
            finally:
                cursor.close()
                connection.close()
            
        finally:
            # Limpiar archivo temporal
            try:
                os.unlink(tmp_path)
            except:
                pass
    
    except Exception as e:
        print(f"[IMPORTAR_PDF] [X] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Error al procesar el PDF: {str(e)}'}), 500


# ============================================================================
# IMPORTAR EXCEL - Extraer materiales y servicios desde un archivo Excel
# ============================================================================

# Palabras clave que indican fila de resumen/totales en Excel (NO son items)
FILAS_RESUMEN_EXCEL = [
    'costo directo', 'costo total', 'gastos generales', 'utilidad',
    'sub total', 'subtotal', 'igv', 'presupuesto total',
    'total de ejecucion', 'total general', 'suma total',
    'gravamen', 'impuesto', 'total materiales', 'total servicios',
    'total parcial', 'costo directo total'
]

# Unidades de tiempo = SERVICIOS (mismo criterio que PDF)
UNIDADES_SERVICIO_EXCEL = [
    # Servicios especiales (HML, ML, GLB)
    'hm', 'ml', 'glb',
    # Tiempo
    'mes', 'meses', 'dia', 'dias', 'da', 'das',
    'semana', 'semanas', 'hr', 'hora', 'horas', 'h',
    'min', 'minuto', 'minutos', 'ao', 'aos',
    'quincena', 'bimestre', 'trimestre', 'semestre'
]

# Unidades que SIEMPRE son materiales (no dependen de la clasificacion)
UNIDADES_MATERIAL_EXCEL = ['und', 'vol', 'bulto', 'm2', 'm3', 'kg', 'lt', 'gl', 'unid', 'par', 'jgo']


def es_codigo_item(valor):
    """Verifica si un valor parece codigo de item (01.01.01 o similar)"""
    if valor is None:
        return False
    texto = str(valor).strip()
    return bool(re.match(r'^\d{2}\.\d{2}\.\d{2}', texto))


def es_fila_resumen(valor):
    """Verifica si un valor es fila de resumen/totales"""
    if valor is None:
        return False
    texto = str(valor).strip().lower()
    return any(p in texto for p in FILAS_RESUMEN_EXCEL)


def clasificar_unidad_excel(unidad_texto):
    """Clasifica una unidad como MATERIAL o SERVICIO"""
    if not unidad_texto:
        return 'MATERIAL'
    u = str(unidad_texto).strip().lower()
    for serv in UNIDADES_SERVICIO_EXCEL:
        if u == serv:
            return 'SERVICIO'
    return 'MATERIAL'


def es_fila_vacia(fila):
    """Verifica si una fila del Excel esta completamente vacia"""
    for celda in fila:
        if celda is not None and str(celda).strip():
            return False
    return True


def detectar_columnas(ws):
    """
    Detecta automaticamente las columnas del Excel buscando la fila de encabezados.
    Retorna un dict con los indices de cada columna encontrada.
    
    Estrategia: busca filas que contengan palabras clave como:
    - Codigo, Item, Nro, N (para columna de codigo)
    - Descripcion, Material, Servicio, Concepto, Detalle (para descripcion)
    - Unidad, Und, U.M. (para unidad)
    - Cantidad, Cant, Vol, Area (para cantidad)
    - Precio, Costo, Unitario, P.U. (para precio)
    - Subtotal, Total, Importe (para subtotal)
    - Tipo, Clase, Categoria (para tipo material/servicio)
    """
    keywords_codigo = ['codigo', 'cod', 'item', 'nro', 'n°', 'nº', 'numero', 'nro.', 'cod.', 'codIGO']
    keywords_descripcion = ['descripcion', 'descripci', 'material', 'servicio', 'concepto', 'detalle', 'articulo', 'producto', 'nombre', 'glosa']
    keywords_unidad = ['unidad', 'und', 'u.m.', 'um', 'medida', 'unid']
    keywords_cantidad = ['cantidad', 'cant', 'vol', 'area', 'metros', 'cant.', 'volumen']
    keywords_precio = ['precio', 'costo', 'unitario', 'p.u.', 'pu', 'importe unitario', 'costo unitario', 'p. unitario']
    keywords_subtotal = ['subtotal', 'total', 'importe', 'monto', 'parcial', 'costo total', 'sub total']
    keywords_tipo = ['tipo', 'clase', 'categoria', 'cat.', 'clasificacion']
    
    columnas = {
        'codigo': None,
        'descripcion': None,
        'unidad': None,
        'cantidad': None,
        'precio': None,
        'subtotal': None,
        'tipo': None,
        'fila_inicio_datos': None
    }
    
    max_filas = min(ws.max_row or 1, 50)  # Buscar en las primeras 50 filas
    
    for num_fila in range(1, max_filas + 1):
        fila = [ws.cell(row=num_fila, column=c).value for c in range(1, (ws.max_column or 10) + 1)]
        
        if es_fila_vacia(fila):
            continue
        
        # Contar cuantas palabras clave matchean en esta fila
        matches = 0
        for celda in fila:
            if celda is None:
                continue
            texto = str(celda).strip().lower()
            for kw_list in [keywords_codigo, keywords_descripcion, keywords_unidad, keywords_cantidad, keywords_precio, keywords_subtotal]:
                if any(kw in texto for kw in kw_list):
                    matches += 1
                    break
        
        # Si hay 3+ matches, probablemente es la fila de encabezados
        if matches >= 3:
            for idx, celda in enumerate(fila):
                if celda is None:
                    continue
                texto = str(celda).strip().lower()
                
                if columnas['codigo'] is None and any(kw in texto for kw in keywords_codigo):
                    columnas['codigo'] = idx
                elif columnas['descripcion'] is None and any(kw in texto for kw in keywords_descripcion):
                    columnas['descripcion'] = idx
                elif columnas['unidad'] is None and any(kw in texto for kw in keywords_unidad):
                    columnas['unidad'] = idx
                elif columnas['cantidad'] is None and any(kw in texto for kw in keywords_cantidad):
                    columnas['cantidad'] = idx
                elif columnas['precio'] is None and any(kw in texto for kw in keywords_precio):
                    columnas['precio'] = idx
                elif columnas['subtotal'] is None and any(kw in texto for kw in keywords_subtotal):
                    columnas['subtotal'] = idx
                elif columnas['tipo'] is None and any(kw in texto for kw in keywords_tipo):
                    columnas['tipo'] = idx
            
            columnas['fila_inicio_datos'] = num_fila + 1
            print(f"[EXCEL_DETECT] Encabezados detectados en fila {num_fila}: {columnas}")
            break
    
    # Si no encontro encabezados, intentar detectar por estructura de datos
    if columnas['fila_inicio_datos'] is None:
        print("[EXCEL_DETECT] No se detectaron encabezados, intentando deteccion por estructura...")
        for num_fila in range(1, min(10, max_filas + 1)):
            fila = [ws.cell(row=num_fila, column=c).value for c in range(1, (ws.max_column or 10) + 1)]
            # Buscar primera fila con un codigo tipo XX.XX.XX
            for idx, celda in enumerate(fila):
                if es_codigo_item(celda):
                    columnas['codigo'] = idx
                    columnas['fila_inicio_datos'] = num_fila
                    # Inferir otras columnas por posicion
                    cols_con_datos = [i for i, c in enumerate(fila) if c is not None and str(c).strip()]
                    if len(cols_con_datos) >= 4:
                        # Asumir orden comun: codigo, descripcion, unidad, cantidad, precio, subtotal
                        posibles_cols = [c for c in cols_con_datos if c != idx]
                        if columnas['descripcion'] is None and len(posibles_cols) > 0:
                            columnas['descripcion'] = posibles_cols[0]
                        if len(posibles_cols) > 1:
                            # Buscar unidad (texto corto) vs cantidad (numero)
                            for pc in posibles_cols[1:]:
                                val = fila[pc]
                                if isinstance(val, str) and len(val) < 10:
                                    columnas['unidad'] = pc
                                elif isinstance(val, (int, float)):
                                    if columnas['cantidad'] is None:
                                        columnas['cantidad'] = pc
                                    elif columnas['precio'] is None:
                                        columnas['precio'] = pc
                                    elif columnas['subtotal'] is None:
                                        columnas['subtotal'] = pc
                    print(f"[EXCEL_DETECT] Deteccion por estructura en fila {num_fila}: {columnas}")
                    break
            if columnas['fila_inicio_datos']:
                break
    
    return columnas


def extraer_items_del_excel(filepath):
    """
    Extrae items de un archivo Excel con deteccion automatica de columnas.
    Maneja variaciones en el formato:
    - Diferentes ordenes de columnas
    - Encabezados en diferentes idiomas (espanol/ingles)
    - Filas de grupo/seccion mezcladas con items
    - Celdas fusionadas
    - Porcentajes en diferentes ubicaciones
    """
    items = []
    porcentajes = {
        'gastos_generales': 0,
        'utilidad': 0,
        'igv': 18
    }
    
    print(f"[EXTRAER_EXCEL] Iniciando extraccion de: {filepath}")
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # Buscar la hoja correcta (la que tenga mas datos)
    ws = None
    max_items = 0
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        items_count = 0
        for row in range(1, min(100, (sheet.max_row or 0) + 1)):
            for col in range(1, min(20, (sheet.max_column or 0) + 1)):
                val = sheet.cell(row=row, column=col).value
                if es_codigo_item(val):
                    items_count += 1
        if items_count > max_items:
            max_items = items_count
            ws = sheet
    
    if ws is None:
        ws = wb.active
    
    print(f"[EXTRAER_EXCEL] Usando hoja: '{ws.title}' con {ws.max_row} filas x {ws.max_column} columnas")
    
    # PASO 1: Buscar porcentajes en todo el documento
    for row in range(1, (ws.max_row or 0) + 1):
        for col in range(1, (ws.max_column or 0) + 1):
            celda = ws.cell(row=row, column=col).value
            if celda is None:
                continue
            texto = str(celda).lower().strip()
            
            match_gg = re.search(r'gastos generales\s*[\(:]?\s*(\d+(?:\.\d+)?)\s*%', texto)
            if match_gg:
                porcentajes['gastos_generales'] = float(match_gg.group(1))
                print(f"[EXTRAER_EXCEL] Porcentaje GG: {porcentajes['gastos_generales']}%")
            
            match_util = re.search(r'utilidad\s*[\(:]?\s*(\d+(?:\.\d+)?)\s*%', texto)
            if match_util:
                porcentajes['utilidad'] = float(match_util.group(1))
                print(f"[EXTRAER_EXCEL] Porcentaje Utilidad: {porcentajes['utilidad']}%")
            
            match_igv = re.search(r'igv\s*[\(:]?\s*(\d+(?:\.\d+)?)\s*%', texto)
            if match_igv:
                porcentajes['igv'] = float(match_igv.group(1))
                print(f"[EXTRAER_EXCEL] Porcentaje IGV: {porcentajes['igv']}%")
    
    # PASO 2: Detectar columnas
    columnas = detectar_columnas(ws)
    
    if columnas['fila_inicio_datos'] is None:
        print("[EXTRAER_EXCEL] [X] No se pudieron detectar los encabezados")
        return items, porcentajes
    
    # PASO 3: Extraer items
    tiene_codigo = columnas['codigo'] is not None
    tiene_descripcion = columnas['descripcion'] is not None
    tiene_unidad = columnas['unidad'] is not None
    tiene_cantidad = columnas['cantidad'] is not None
    tiene_precio = columnas['precio'] is not None
    tiene_subtotal = columnas['subtotal'] is not None
    tiene_tipo = columnas['tipo'] is not None
    
    print(f"[EXTRAER_EXCEL] Columnas detectadas: codigo={columnas['codigo']}, desc={columnas['descripcion']}, "
          f"unidad={columnas['unidad']}, cant={columnas['cantidad']}, "
          f"precio={columnas['precio']}, sub={columnas['subtotal']}, tipo={columnas['tipo']}")
    
    for num_fila in range(columnas['fila_inicio_datos'], (ws.max_row or 0) + 1):
        # Leer toda la fila
        max_col = ws.max_column or 10
        fila = [ws.cell(row=num_fila, column=c).value for c in range(1, max_col + 1)]
        
        # Saltar filas vacias
        if es_fila_vacia(fila):
            continue
        
        # Obtener valores de cada columna
        codigo_raw = fila[columnas['codigo']] if tiene_codigo and columnas['codigo'] < len(fila) else None
        desc_raw = fila[columnas['descripcion']] if tiene_descripcion and columnas['descripcion'] < len(fila) else None
        unidad_raw = fila[columnas['unidad']] if tiene_unidad and columnas['unidad'] < len(fila) else None
        cant_raw = fila[columnas['cantidad']] if tiene_cantidad and columnas['cantidad'] < len(fila) else None
        precio_raw = fila[columnas['precio']] if tiene_precio and columnas['precio'] < len(fila) else None
        subtotal_raw = fila[columnas['subtotal']] if tiene_subtotal and columnas['subtotal'] < len(fila) else None
        tipo_raw = fila[columnas['tipo']] if tiene_tipo and columnas['tipo'] < len(fila) else None
        
        # Si no tiene columna de descripcion, buscar en todas las celdas de texto
        if desc_raw is None:
            for celda in fila:
                if celda is not None and isinstance(celda, str) and len(celda) > 5:
                    desc_raw = celda
                    break
        
        # Si no tiene codigo, intentar extraer de la primera celda con formato XX.XX.XX
        if codigo_raw is None and not tiene_codigo:
            for celda in fila:
                if es_codigo_item(celda):
                    codigo_raw = celda
                    break
        
        # Filtrar filas de resumen
        if desc_raw and es_fila_resumen(desc_raw):
            print(f"[EXTRAER_EXCEL] Fila {num_fila} descartada (resumen): {desc_raw}")
            continue
        
        # Si no hay descripcion ni codigo, saltar
        if desc_raw is None and codigo_raw is None:
            continue
        
        # Convertir valores
        codigo = str(codigo_raw).strip() if codigo_raw else ''
        descripcion = str(desc_raw).strip() if desc_raw else ''
        unidad = str(unidad_raw).strip() if unidad_raw else ''
        
        # Convertir cantidad
        cantidad = 0
        if cant_raw is not None:
            try:
                cantidad = float(Decimal(str(cant_raw).replace(',', '')).normalize())
            except (ValueError, TypeError, InvalidOperation):
                cantidad = 0
        
        # Convertir precio
        precio = 0
        if precio_raw is not None:
            try:
                precio = float(Decimal(str(precio_raw).replace(',', '')).normalize())
            except (ValueError, TypeError, InvalidOperation):
                precio = 0
        
        # Convertir subtotal
        subtotal = 0
        if subtotal_raw is not None:
            try:
                subtotal = float(Decimal(str(subtotal_raw).replace(',', '')).normalize())
            except (ValueError, TypeError, InvalidOperation):
                subtotal = 0
        
        # Si no hay subtotal pero hay cantidad y precio, calcularlo
        if subtotal == 0 and cantidad > 0 and precio > 0:
            subtotal = cantidad * precio
        
        # Si cantidad es 0, intentar buscar en otras columnas numericas
        if cantidad == 0 and precio == 0:
            for celda in fila:
                if isinstance(celda, (int, float)) and celda > 0:
                    if cantidad == 0:
                        cantidad = celda
                    elif precio == 0:
                        precio = celda
            if subtotal == 0 and cantidad > 0 and precio > 0:
                subtotal = cantidad * precio
        
        # Saltar si no tiene datos utiles
        if cantidad <= 0 and precio <= 0:
            print(f"[EXTRAER_EXCEL] Fila {num_fila} descartada (sin datos numericos): {descripcion[:50]}")
            continue
        
        # Si no tiene unidad, es titulo/subtitulo (no material ni servicio)
        if not unidad:
            print(f"[EXTRAER_EXCEL] Fila {num_fila} descartada (sin unidad): {descripcion[:50]}")
            continue
        
        # Clasificar tipo
        tipo = clasificar_unidad_excel(unidad)
        
        items.append({
            'codigo': codigo,
            'descripcion': descripcion,
            'cantidad': cantidad,
            'precio_unitario': precio,
            'subtotal': subtotal,
            'unidad': unidad,
            'tipo': tipo
        })
        
        print(f"[EXTRAER_EXCEL] [OK] Fila {num_fila}: {descripcion[:40]} | {unidad} | {cantidad} | {precio} | {subtotal} | {tipo}")
    
    print(f"[EXTRAER_EXCEL] Total: {len(items)} items extraidos")
    return items, porcentajes


@main_bp.route('/api/presupuestos/importar-excel', methods=['POST'])
def importar_presupuesto_excel():
    """
    Recibe un Excel, extrae items, clasifica materiales/servicios.
    Crea automaticamente los materiales que no existen en TblMateriales.
    Mismo comportamiento que importar-pdf pero para archivos .xlsx/.xls
    """
    try:
        if 'archivo' not in request.files:
            return jsonify({'success': False, 'error': 'No se proporcion ningun archivo'}), 400
        
        archivo = request.files['archivo']
        
        if not archivo.filename:
            return jsonify({'success': False, 'error': 'Nombre de archivo vacio'}), 400
        
        nombre_lower = archivo.filename.lower()
        if not nombre_lower.endswith('.xlsx') and not nombre_lower.endswith('.xls'):
            return jsonify({'success': False, 'error': 'El archivo debe ser un Excel (.xlsx o .xls)'}), 400
        
        # Guardar temporalmente
        suffix = '.xlsx' if nombre_lower.endswith('.xlsx') else '.xls'
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            archivo.save(tmp.name)
            tmp_path = tmp.name
        
        try:
            # Extraer items del Excel y porcentajes
            items, porcentajes_excel = extraer_items_del_excel(tmp_path)
            
            if not items:
                return jsonify({
                    'success': False,
                    'error': 'No se encontraron items en el Excel. Verifica que el documento contenga columnas con: codigo, descripcion, unidad, cantidad y precio.'
                }), 400
            
            # Conexion a BD para buscar/crear materiales
            connection = get_db_connection()
            if not connection:
                return jsonify({'success': False, 'error': 'Error de conexion a la base de datos'}), 500
            
            try:
                cursor = connection.cursor(dictionary=True)
                
                # Separar materiales y servicios
                materiales = []
                servicios = []
                materiales_nuevos = []
                materiales_existentes = []
                materiales_error = []
                
                print(f"[IMPORTAR_EXCEL] Procesando {len(items)} items extraidos del Excel")
                
                for idx, item in enumerate(items):
                    print(f"[IMPORTAR_EXCEL] Item {idx + 1}/{len(items)}: {item['descripcion'][:40]}... tipo={item['tipo']}")
                    
                    if item['tipo'] == 'SERVICIO':
                        servicios.append({
                            'descripcion': item['descripcion'],
                            'cantidad': item['cantidad'],
                            'precio_unitario': item['precio_unitario'],
                            'subtotal': item['subtotal'],
                            'unidad': item['unidad']
                        })
                        print(f"[IMPORTAR_EXCEL]    Servicio agregado")
                    else:
                        nombre_material = item['descripcion']
                        id_material = None
                        
                        # 1. Buscar si el material ya existe (nombre exacto)
                        cursor.execute(
                            "SELECT id_material FROM TblMateriales WHERE nombre = %s AND estado = 'ACTIVO' LIMIT 1",
                            (nombre_material,)
                        )
                        existente = cursor.fetchone()
                        
                        if existente:
                            id_material = existente['id_material']
                            materiales_existentes.append(nombre_material)
                            print(f"[IMPORTAR_EXCEL]    Material ya existe: ID={id_material}")
                        else:
                            # 2. Buscar id_unidad por abreviatura
                            unidad_excel = item['unidad']
                            id_unidad = None
                            
                            if unidad_excel:
                                cursor.execute(
                                    """SELECT id_unidad FROM TblUnidadMedida
                                       WHERE (abreviatura = %s OR nombre = %s OR nombre LIKE %s OR abreviatura LIKE %s)
                                       AND estado = 'ACTIVO' LIMIT 1""",
                                    (unidad_excel, unidad_excel, f'%{unidad_excel}%', f'%{unidad_excel}%')
                                )
                                unidad_result = cursor.fetchone()
                                
                                if unidad_result:
                                    id_unidad = unidad_result['id_unidad']
                                    print(f"[IMPORTAR_EXCEL]    Unidad encontrada: {unidad_excel} -> id_unidad={id_unidad}")
                                else:
                                    id_unidad = 1
                                    print(f"[IMPORTAR_EXCEL]    Unidad '{unidad_excel}' no encontrada, usando UND=1")
                            else:
                                id_unidad = 1
                            
                            # 3. Crear material usando SP
                            cursor.execute("SET @p_id = NULL, @p_codigo = NULL, @p_result = 0")
                            cursor.execute(
                                """CALL sp_CrearMaterialConCodigoAuto(
                                    %s, %s, NULL, %s, NULL,
                                    @p_id, @p_codigo, @p_result
                                )""",
                                (nombre_material, item.get('descripcion', nombre_material), id_unidad)
                            )
                            cursor.execute("SELECT @p_id as id, @p_codigo as codigo, @p_result as resultado")
                            sp_result = cursor.fetchone()
                            
                            if sp_result and sp_result['resultado'] == 1:
                                id_material = sp_result['id']
                                materiales_nuevos.append(nombre_material)
                                print(f"[IMPORTAR_EXCEL]    Material CREADO: {sp_result['codigo']} (ID={id_material})")
                            else:
                                materiales_error.append(nombre_material)
                                print(f"[IMPORTAR_EXCEL]    ERROR creando material: resultado={sp_result}")
                        
                        materiales.append({
                            'id_material': id_material,
                            'nombre': nombre_material,
                            'cantidad': item['cantidad'],
                            'precio_unitario': item['precio_unitario'],
                            'subtotal': item['subtotal'],
                            'unidad': item['unidad'],
                            'es_nuevo': id_material is not None and nombre_material in materiales_nuevos
                        })
                
                print(f"[IMPORTAR_EXCEL] [OK] Excel procesado: {len(materiales)} materiales ({len(materiales_nuevos)} nuevos, {len(materiales_existentes)} existentes, {len(materiales_error)} errores), {len(servicios)} servicios")
                
                return jsonify({
                    'success': True,
                    'materiales': materiales,
                    'servicios': servicios,
                    'total_items': len(items),
                    'materiales_nuevos': materiales_nuevos,
                    'materiales_existentes': materiales_existentes,
                    'materiales_error': materiales_error,
                    'porcentajes': porcentajes_excel,
                    'mensaje': f'Se extrajeron {len(materiales)} materiales ({len(materiales_nuevos)} nuevos, {len(materiales_existentes)} existentes) y {len(servicios)} servicios'
                })
                
            finally:
                cursor.close()
                connection.close()
            
        finally:
            try:
                os.unlink(tmp_path)
            except:
                pass
    
    except Exception as e:
        print(f"[IMPORTAR_EXCEL] [X] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Error al procesar el Excel: {str(e)}'}), 500


# ============================================================================
# API: SISTEMA DE VERSIONES DE PRESUPUESTO
# ============================================================================

@main_bp.route('/api/presupuestos/guardar-snapshot', methods=['POST'])
@login_required
def guardar_snapshot_presupuesto():
    """
    Guardar snapshot del presupuesto actual ANTES de editar.
    Se ejecuta automáticamente al hacer clic en "Editar"
    """
    try:
        data = request.get_json()
        id_presupuesto = data.get('id_presupuesto')
        user_documento = session.get('user_documento')
        
        if not id_presupuesto:
            return jsonify({'success': False, 'error': 'ID de presupuesto requerido'}), 400
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Error de conexión'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            print(f"[VERSIONES] 📸 Guardando snapshot de presupuesto {id_presupuesto}")
            
            # Llamar SP para guardar snapshot
            cursor.callproc('sp_GuardarSnapshotAntesDeEditar', (id_presupuesto, user_documento))
            
            # ✅ Consumir resultados (aunque el SP ya no retorna SELECT)
            for res in cursor.stored_results():
                res.fetchall()  # Consumir cualquier resultado pendiente
            
            connection.commit()
            cursor.close()
            connection.close()
            
            # ✅ El SP no retorna datos, pero si llegamos aquí, se ejecutó exitosamente
            print(f"[VERSIONES] ✅ Snapshot guardado exitosamente")
            return jsonify({
                'success': True,
                'mensaje': 'Snapshot guardado correctamente'
            }), 200
                
        except Error as e:
            print(f"[VERSIONES] ❌ Error SQL: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
            
    except Exception as e:
        print(f"[VERSIONES] ❌ Error general: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@main_bp.route('/api/presupuestos/<int:id_presupuesto>/versiones', methods=['GET'])
@login_required
def obtener_historial_versiones(id_presupuesto):
    """Obtener historial de versiones de un presupuesto"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Error de conexión'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            print(f"[VERSIONES] 📋 Obteniendo historial de presupuesto {id_presupuesto}")
            
            # Llamar SP para obtener historial
            cursor.callproc('sp_ObtenerHistorialVersiones', (id_presupuesto,))
            
            versiones = []
            for result in cursor.stored_results():
                versiones = result.fetchall()
                break
            
            # Serializar fechas
            for version in versiones:
                if version.get('fecha_version'):
                    version['fecha_version'] = version['fecha_version'].isoformat()
            
            cursor.close()
            connection.close()
            
            print(f"[VERSIONES] ✅ {len(versiones)} versiones encontradas")
            
            return jsonify({
                'success': True,
                'versiones': versiones
            }), 200
                
        except Error as e:
            print(f"[VERSIONES] ❌ Error SQL: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
            
    except Exception as e:
        print(f"[VERSIONES] ❌ Error general: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@main_bp.route('/api/presupuestos/<int:id_presupuesto>/versiones/<int:numero_version>', methods=['GET'])
@login_required
def obtener_version_especifica(id_presupuesto, numero_version):
    """Obtener una versión específica del presupuesto con su detalle"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Error de conexión'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            print(f"[VERSIONES] 📄 Obteniendo versión {numero_version} de presupuesto {id_presupuesto}")
            
            # Llamar SP para obtener versión específica
            cursor.callproc('sp_ObtenerVersionPresupuesto', (id_presupuesto, numero_version))
            
            # Primer resultset: datos de la versión
            version_data = None
            for result in cursor.stored_results():
                version_data = result.fetchone()
                break
            
            # Segundo resultset: detalle de partidas
            cursor.callproc('sp_ObtenerVersionPresupuesto', (id_presupuesto, numero_version))
            partidas = []
            result_count = 0
            for result in cursor.stored_results():
                if result_count == 1:  # Segundo resultset
                    partidas = result.fetchall()
                result_count += 1
            
            cursor.close()
            connection.close()
            
            if not version_data:
                return jsonify({'success': False, 'error': 'Versión no encontrada'}), 404
            
            # Serializar fechas
            for key, value in version_data.items():
                if isinstance(value, datetime):
                    version_data[key] = value.isoformat()
                elif hasattr(value, 'isoformat'):
                    version_data[key] = value.isoformat()
            
            for partida in partidas:
                for key, value in partida.items():
                    if isinstance(value, Decimal):
                        partida[key] = float(value)
            
            print(f"[VERSIONES] ✅ Versión {numero_version} obtenida con {len(partidas)} partidas")
            
            return jsonify({
                'success': True,
                'version': version_data,
                'partidas': partidas
            }), 200
                
        except Error as e:
            print(f"[VERSIONES] ❌ Error SQL: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
            
    except Exception as e:
        print(f"[VERSIONES] ❌ Error general: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@main_bp.route('/api/presupuestos/<int:id_presupuesto>/version-actual', methods=['GET'])
@login_required
def obtener_version_actual(id_presupuesto):
    """Obtener número de versión actual del presupuesto"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Error de conexión'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            # Llamar SP para obtener versión actual
            cursor.callproc('sp_ObtenerVersionActual', (id_presupuesto,))
            
            result = None
            for res in cursor.stored_results():
                result = res.fetchone()
                break
            
            cursor.close()
            connection.close()
            
            return jsonify({
                'success': True,
                'version_actual': result.get('version_actual', 0) if result else 0,
                'total_versiones': result.get('total_versiones', 0) if result else 0
            }), 200
                
        except Error as e:
            print(f"[VERSIONES] ❌ Error SQL: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
            
    except Exception as e:
        print(f"[VERSIONES] ❌ Error general: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@main_bp.route('/api/presupuestos/comparar-versiones', methods=['POST'])
@login_required
def comparar_versiones():
    """Comparar dos versiones de un presupuesto"""
    try:
        data = request.get_json()
        id_presupuesto = data.get('id_presupuesto')
        version1 = data.get('version1')
        version2 = data.get('version2')
        
        if not all([id_presupuesto, version1, version2]):
            return jsonify({'success': False, 'error': 'Faltan parámetros requeridos'}), 400
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Error de conexión'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            print(f"[VERSIONES] 🔄 Comparando v{version1} vs v{version2} de presupuesto {id_presupuesto}")
            
            # Llamar SP para comparar versiones
            cursor.callproc('sp_CompararVersionesPresupuesto', (id_presupuesto, version1, version2))
            
            comparacion = None
            for result in cursor.stored_results():
                comparacion = result.fetchone()
                break
            
            cursor.close()
            connection.close()
            
            if not comparacion:
                return jsonify({'success': False, 'error': 'No se pudo comparar las versiones'}), 500
            
            # Serializar fechas
            for key, value in comparacion.items():
                if isinstance(value, datetime):
                    comparacion[key] = value.isoformat()
                elif isinstance(value, Decimal):
                    comparacion[key] = float(value)
            
            print(f"[VERSIONES] ✅ Comparación realizada")
            
            return jsonify({
                'success': True,
                'comparacion': comparacion
            }), 200
                
        except Error as e:
            print(f"[VERSIONES] ❌ Error SQL: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
            
    except Exception as e:
        print(f"[VERSIONES] ❌ Error general: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
