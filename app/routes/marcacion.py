"""
Rutas para el sistema de marcacin de asistencia
"""
from flask import Blueprint, render_template, request, session, jsonify, send_file
from functools import wraps
from mysql.connector import Error
from app.config import DatabaseConfig
from datetime import datetime
import mysql.connector
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

# Crear blueprint para marcacin
marcacion_bp = Blueprint('marcacion', __name__)

# ============================================================================
# UTILIDADES
# ============================================================================

def get_db_connection():
    """Crear conexin a la base de datos Kallpa"""
    try:
        params = DatabaseConfig.get_connection_params()
        connection = mysql.connector.connect(**params)
        return connection
    except Error as e:
        print(f"Error de conexin: {e}")
        return None


def login_required(f):
    """Decorador para proteger rutas que requieren autenticacin"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_documento' not in session and 'user_email' not in session:
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return {'success': False, 'message': 'No autenticado'}, 401
            from flask import redirect, url_for, flash
            flash('Debes iniciar sesin para acceder a esta pgina', 'warning')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function


# ============================================================================
# RUTA: PGINA DE REPORTE DE ASISTENCIA (ADMIN)
# ============================================================================

@marcacion_bp.route('/marcacion')
@login_required
def reporte_asistencia():
    """Pgina de reporte de asistencia de todos los usuarios (vista administrativa)"""
    return render_template('reporte_asistencia.html')


# ============================================================================
# RUTA: PGINA DE CONTROL DE ASISTENCIA (ADMIN)
# ============================================================================

@marcacion_bp.route('/reportes')
@login_required
def reportes():
    """Pgina de reportes de asistencia"""
    return render_template('control_asistencia.html')


# ============================================================================
# RUTA: PGINA DE MARCACIN INDIVIDUAL
# ============================================================================

@marcacion_bp.route('/marcacion-kallpa')
@login_required
def marcacion_kallpa():
    """Pgina de marcacin de Kallpa (mvil)"""
    return render_template('marcacion_kallpa.html')


# ============================================================================
# API: REGISTRAR MARCACIN
# ============================================================================

@marcacion_bp.route('/api/marcacion/registrar', methods=['POST'])
@login_required
def registrar_marcacion():
    """Registrar marcacin (entrada/salida) con GPS y foto"""
    from flask import current_app
    
    try:
        data = request.get_json()
        num_documento = session.get('user_documento')
        tipo_marcacion = data.get('tipo_marcacion')  # 'ENTRADA' o 'SALIDA'
        latitud = data.get('latitud')
        longitud = data.get('longitud')
        precision = data.get('precision')
        foto_base64 = data.get('foto_base64')
        
        print(f"[MARCACION] [...] Intentando registrar: documento={num_documento}, tipo={tipo_marcacion}")
        print(f"[MARCACION]  GPS: lat={latitud}, lon={longitud}, precisin={precision}")
        print(f"[MARCACION]  Foto: {'S' if foto_base64 else 'No'}")
        
        if not num_documento or not tipo_marcacion:
            print(f"[MARCACION] [X] Datos incompletos")
            return jsonify({'success': False, 'error': 'Datos incompletos'}), 400
        
        connection = get_db_connection()
        if not connection:
            print(f"[MARCACION] [X] Error de conexin a BD")
            return jsonify({'success': False, 'error': 'Error de conexin'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            print(f"[MARCACION]  Llamando a SP con GPS y foto")
            
            # Llamar SP para registrar marcacin con GPS y foto
            cursor.execute("""
                CALL sp_RegistrarMarcacionCompleta(
                    %s,  -- p_num_documento
                    %s,  -- p_tipo_marcacion
                    %s,  -- p_latitud
                    %s,  -- p_longitud
                    %s,  -- p_precision
                    %s,  -- p_foto_base64
                    @p_id_marcacion,
                    @p_mensaje
                )
            """, (num_documento, tipo_marcacion, latitud, longitud, precision, foto_base64))
            
            # Leer OUT parameters
            cursor.execute("SELECT @p_id_marcacion as id_marcacion, @p_mensaje as mensaje")
            result = cursor.fetchone()
            
            print(f"[MARCACION] [-] Resultado SP: {result}")
            
            # [OK] COMMIT CRTICO: Guardar cambios en la BD
            connection.commit()
            print(f"[MARCACION]  COMMIT realizado")
            
            cursor.close()
            connection.close()
            
            if result and result.get('id_marcacion', 0) > 0:
                print(f"[MARCACION] [OK] xito: ID={result['id_marcacion']}, Mensaje={result.get('mensaje')}")
                return jsonify({
                    'success': True,
                    'message': result.get('mensaje', 'Marcacin registrada exitosamente'),
                    'id_marcacion': result['id_marcacion']
                }), 201
            else:
                print(f"[MARCACION] [!] SP retorn ID=0 o NULL")
                return jsonify({
                    'success': False,
                    'error': result.get('mensaje', 'Error al registrar marcacin')
                }), 400
        
        except Error as e:
            print(f"[MARCACION] [X] Error SQL: {str(e)}")
            current_app.logger.error(f"Error SQL en marcacin: {str(e)}")
            return jsonify({'success': False, 'error': f'Error en la base de datos: {str(e)}'}), 500
    
    except Exception as e:
        print(f"[MARCACION] [X] Error general: {str(e)}")
        current_app.logger.error(f"Error general en marcacin: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'error': 'Error en el servidor'}), 500


# ============================================================================
# API: OBTENER ESTADO DE MARCACIN
# ============================================================================

@marcacion_bp.route('/api/marcacion/estado', methods=['GET'])
@login_required
def obtener_estado_marcacion():
    """Obtener el estado actual de marcacin del usuario (ltima entrada sin salida)"""
    from flask import current_app
    
    try:
        num_documento = session.get('user_documento')
        
        if not num_documento:
            return jsonify({'success': False, 'error': 'Usuario no identificado'}), 401
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Error de conexin'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            # Llamar SP para obtener el estado de marcacin
            cursor.callproc('sp_ObtenerEstadoMarcacion', (num_documento,))
            
            # Obtener resultado
            ultima_marcacion = None
            for result in cursor.stored_results():
                result_list = result.fetchall()
                if result_list:
                    ultima_marcacion = result_list[0]
            
            cursor.close()
            connection.close()
            
            # Determinar el estado segn la ltima marcacin
            if ultima_marcacion:
                if ultima_marcacion['tipo_marcacion'] == 'ENTRADA':
                    estado = 'DENTRO'
                else:
                    estado = 'FUERA'
                
                # Extraer hora
                hora = ultima_marcacion.get('hora')
                if isinstance(hora, datetime):
                    hora_str = hora.strftime('%H:%M:%S')
                elif hasattr(hora, 'total_seconds'):  # timedelta
                    total_seconds = int(hora.total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    seconds = total_seconds % 60
                    hora_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                else:
                    hora_str = str(hora) if hora else None
                
                return jsonify({
                    'success': True,
                    'estado': estado,
                    'ultima_marcacion': ultima_marcacion['tipo_marcacion'],
                    'hora': hora_str
                }), 200
            else:
                return jsonify({
                    'success': True,
                    'estado': 'FUERA',
                    'ultima_marcacion': None,
                    'hora': None
                }), 200
        
        except Error as e:
            current_app.logger.error(f"Error SQL en estado marcacin: {str(e)}")
            return jsonify({'success': False, 'error': 'Error en la base de datos'}), 500
    
    except Exception as e:
        current_app.logger.error(f"Error general en estado marcacin: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'error': 'Error en el servidor'}), 500


# ============================================================================
# API: OBTENER HISTORIAL DE MARCACIN
# ============================================================================

@marcacion_bp.route('/api/marcacion/historial', methods=['GET'])
@login_required
def obtener_historial_marcacion():
    """Obtener historial de marcaciones del usuario con estados de asistencia"""
    try:
        num_documento = session.get('user_documento')
        dias = request.args.get('dias', 7, type=int)
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Error de conexin'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            print(f"[HISTORIAL] [...] Obteniendo historial CON ESTADOS para documento {num_documento}, ltimos {dias} das")
            
            # Llamar SP que calcula estados de asistencia
            cursor.callproc('sp_ObtenerHistorialConEstados', (num_documento, dias))
            
            # Obtener resultados
            marcaciones = []
            for result in cursor.stored_results():
                marcaciones = result.fetchall()
            
            print(f"[HISTORIAL] [OK] {len(marcaciones)} marcaciones encontradas con estados")
            
            #  Serializar datos para JSON (convertir datetime, date, timedelta, etc.)
            import datetime as dt_module
            for marcacion in marcaciones:
                for key, value in marcacion.items():
                    # Convertir datetime a ISO string
                    if isinstance(value, datetime):
                        marcacion[key] = value.isoformat()
                    # Convertir date a string YYYY-MM-DD
                    elif isinstance(value, dt_module.date):
                        marcacion[key] = value.strftime('%Y-%m-%d')
                    # Convertir timedelta a string
                    elif hasattr(value, 'total_seconds'):  # timedelta
                        marcacion[key] = str(value)
                    # Asegurar que cualquier fecha sea string
                    elif hasattr(value, 'strftime'):
                        marcacion[key] = value.strftime('%Y-%m-%d')
            
            cursor.close()
            connection.close()
            
            return jsonify({
                'success': True,
                'data': marcaciones
            }), 200
        
        except Error as e:
            print(f"[HISTORIAL] [X] Error SQL: {e}")
            return jsonify({'success': False, 'error': 'Error en la base de datos'}), 500
    
    except Exception as e:
        print(f"[HISTORIAL] [X] Error general: {e}")
        return jsonify({'success': False, 'error': 'Error del servidor'}), 500


# ============================================================================
# API: OBTENER REPORTE DE ASISTENCIA DE TODOS LOS USUARIOS
# ============================================================================

@marcacion_bp.route('/api/marcacion/reporte-todos', methods=['GET'])
@login_required
def obtener_reporte_todos():
    """Obtener reporte de asistencia de todos los usuarios"""
    try:
        # Parmetros de filtro
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        num_documento = request.args.get('num_documento')  # Filtro opcional por usuario
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Error de conexin'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            print(f"[REPORTE_TODOS] [...] Obteniendo reporte de asistencia")
            print(f"[REPORTE_TODOS]  Fechas: desde={fecha_desde}, hasta={fecha_hasta}")
            
            # Llamar SP para obtener reporte de marcaciones
            cursor.callproc('sp_ReporteMarcacionesTodos', (
                fecha_desde if fecha_desde else None,
                fecha_hasta if fecha_hasta else None,
                num_documento if num_documento else None
            ))
            
            # Obtener resultados
            marcaciones = []
            for result in cursor.stored_results():
                marcaciones = result.fetchall()
            
            print(f"[REPORTE_TODOS] [OK] {len(marcaciones)} marcaciones encontradas")
            
            # Serializar datos para JSON
            import datetime as dt_module
            for marcacion in marcaciones:
                for key, value in marcacion.items():
                    if isinstance(value, datetime):
                        marcacion[key] = value.isoformat()
                    elif isinstance(value, dt_module.date):
                        marcacion[key] = value.strftime('%Y-%m-%d')
                    elif isinstance(value, dt_module.time):
                        marcacion[key] = value.strftime('%H:%M:%S')
                    elif hasattr(value, 'total_seconds'):
                        marcacion[key] = str(value)
            
            cursor.close()
            connection.close()
            
            return jsonify({
                'success': True,
                'data': marcaciones
            }), 200
        
        except Error as e:
            print(f"[REPORTE_TODOS] [X] Error SQL: {e}")
            return jsonify({'success': False, 'error': 'Error en la base de datos'}), 500
    
    except Exception as e:
        print(f"[REPORTE_TODOS] [X] Error general: {e}")
        return jsonify({'success': False, 'error': 'Error del servidor'}), 500


# ============================================================================
# API: OBTENER RESUMEN DE ASISTENCIA POR USUARIO
# ============================================================================

@marcacion_bp.route('/api/marcacion/resumen-usuarios', methods=['GET'])
@login_required
def obtener_resumen_usuarios():
    """Obtener resumen de asistencia por usuario (das trabajados, puntualidad, etc.)"""
    try:
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Error de conexin'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            print(f"[RESUMEN_USUARIOS] [...] Obteniendo resumen de usuarios")
            print(f"[RESUMEN_USUARIOS]  Fechas: desde={fecha_desde}, hasta={fecha_hasta}")
            
            # Llamar SP para obtener resumen de usuarios
            cursor.callproc('sp_ResumenUsuariosAsistencia', (
                fecha_desde if fecha_desde else None,
                fecha_hasta if fecha_hasta else None
            ))
            
            # Obtener resultados
            resumen = []
            for result in cursor.stored_results():
                resumen = result.fetchall()
            
            print(f"[RESUMEN_USUARIOS] [OK] {len(resumen)} usuarios en el resumen")
            
            # Serializar datos
            import datetime as dt_module
            for item in resumen:
                for key, value in item.items():
                    if isinstance(value, datetime):
                        item[key] = value.isoformat()
                    elif isinstance(value, dt_module.date):
                        item[key] = value.strftime('%Y-%m-%d')
            
            cursor.close()
            connection.close()
            
            return jsonify({
                'success': True,
                'data': resumen
            }), 200
        
        except Error as e:
            print(f"[RESUMEN_USUARIOS] [X] Error SQL: {e}")
            return jsonify({'success': False, 'error': 'Error en la base de datos'}), 500
    
    except Exception as e:
        print(f"[RESUMEN_USUARIOS] [X] Error general: {e}")
        return jsonify({'success': False, 'error': 'Error del servidor'}), 500


# ============================================================================
# API: OBTENER DETALLE DE MARCACIONES DE UN DA ESPECFICO
# ============================================================================

@marcacion_bp.route('/api/marcacion/detalle-dia', methods=['GET'])
@login_required
def obtener_detalle_dia():
    """Obtener todas las marcaciones de un usuario en un da especfico con fotos y GPS"""
    try:
        num_documento = request.args.get('num_documento')
        fecha = request.args.get('fecha')
        
        if not num_documento or not fecha:
            return jsonify({'success': False, 'error': 'Faltan parmetros requeridos'}), 400
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Error de conexin'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            print(f"[DETALLE_DIA] [...] Obteniendo detalle para documento={num_documento}, fecha={fecha}")
            
            # Llamar SP para obtener detalle del da
            cursor.callproc('sp_ObtenerDetalleDia', (num_documento, fecha))
            
            # Obtener ambos resultsets
            usuario = None
            marcaciones = []
            
            result_index = 0
            for result in cursor.stored_results():
                result_data = result.fetchall()
                if result_index == 0:
                    # Primer resultset: informacin del usuario
                    if result_data:
                        usuario = result_data[0]
                elif result_index == 1:
                    # Segundo resultset: marcaciones del da
                    marcaciones = result_data
                result_index += 1
            
            if not usuario:
                cursor.close()
                connection.close()
                return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404
            
            print(f"[DETALLE_DIA] [OK] {len(marcaciones)} marcaciones encontradas")
            
            # Serializar datos
            import datetime as dt_module
            for marcacion in marcaciones:
                for key, value in marcacion.items():
                    if isinstance(value, datetime):
                        marcacion[key] = value.isoformat()
                    elif isinstance(value, dt_module.time):
                        marcacion[key] = value.strftime('%H:%M:%S')
                    elif isinstance(value, dt_module.date):
                        marcacion[key] = value.strftime('%Y-%m-%d')
                    elif isinstance(value, dt_module.timedelta):
                        # Convertir timedelta a string de tiempo HH:MM:SS
                        total_seconds = int(value.total_seconds())
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        seconds = total_seconds % 60
                        marcacion[key] = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            
            cursor.close()
            connection.close()
            
            return jsonify({
                'success': True,
                'usuario': usuario,
                'fecha': fecha,
                'marcaciones': marcaciones,
                'total_marcaciones': len(marcaciones)
            }), 200
        
        except Error as e:
            print(f"[DETALLE_DIA] [X] Error SQL: {e}")
            return jsonify({'success': False, 'error': 'Error en la base de datos'}), 500
    
    except Exception as e:
        print(f"[DETALLE_DIA] [X] Error general: {e}")
        return jsonify({'success': False, 'error': 'Error del servidor'}), 500


# ============================================================================
# API: EXPORTAR MARCACIONES A EXCEL
# ============================================================================

@marcacion_bp.route('/api/marcacion/exportar-excel', methods=['GET'])
@login_required
def exportar_marcaciones_excel():
    """Exportar marcaciones detalladas a Excel con turnos organizados"""
    try:
        # Obtener parmetros de filtro
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        
        if not fecha_desde or not fecha_hasta:
            return jsonify({'success': False, 'error': 'Fechas requeridas'}), 400
        
        print(f"[EXPORTAR_EXCEL] [...] Exportando marcaciones: {fecha_desde} a {fecha_hasta}")
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Error de conexin'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            # Llamar SP para obtener datos detallados
            cursor.callproc('sp_ExportarMarcacionDetallada', (fecha_desde, fecha_hasta))
            
            # Obtener resultados
            datos = []
            for result in cursor.stored_results():
                datos = result.fetchall()
            
            print(f"[EXPORTAR_EXCEL] [OK] {len(datos)} registros encontrados")
            
            cursor.close()
            connection.close()
            
            if not datos:
                return jsonify({'success': False, 'error': 'No hay datos para exportar'}), 404
            
            # ====================================================================
            # CREAR ARCHIVO EXCEL
            # ====================================================================
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Marcaciones"
            
            # Estilos
            header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            cell_alignment = Alignment(horizontal="left", vertical="center")
            cell_alignment_center = Alignment(horizontal="center", vertical="center")
            
            border_style = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # ====================================================================
            # ENCABEZADOS
            # ====================================================================
            
            headers = [
                'Nmero Documento',
                'Nombres Completos',
                'Fecha',
                'Entrada 1',
                'Salida 1',
                'Entrada 2',
                'Salida 2',
                'Estado'
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border_style
            
            # ====================================================================
            # DATOS
            # ====================================================================
            
            row_num = 2
            for registro in datos:
                # Nmero Documento
                cell = ws.cell(row=row_num, column=1)
                cell.value = registro.get('Nmero Documento')
                cell.alignment = cell_alignment_center
                cell.border = border_style
                
                # Nombres Completos
                cell = ws.cell(row=row_num, column=2)
                cell.value = registro.get('Nombres Completos')
                cell.alignment = cell_alignment
                cell.border = border_style
                
                # Fecha
                cell = ws.cell(row=row_num, column=3)
                fecha_val = registro.get('Fecha')
                if isinstance(fecha_val, datetime):
                    cell.value = fecha_val.strftime('%Y-%m-%d')
                elif hasattr(fecha_val, 'strftime'):
                    cell.value = fecha_val.strftime('%Y-%m-%d')
                else:
                    cell.value = str(fecha_val) if fecha_val else ''
                cell.alignment = cell_alignment_center
                cell.border = border_style
                
                # Entrada 1
                cell = ws.cell(row=row_num, column=4)
                entrada1 = registro.get('Entrada 1')
                if entrada1:
                    if hasattr(entrada1, 'total_seconds'):  # timedelta
                        total_seconds = int(entrada1.total_seconds())
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        seconds = total_seconds % 60
                        cell.value = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                    else:
                        cell.value = str(entrada1)
                else:
                    cell.value = ''
                cell.alignment = cell_alignment_center
                cell.border = border_style
                
                # Salida 1
                cell = ws.cell(row=row_num, column=5)
                salida1 = registro.get('Salida 1')
                if salida1:
                    if hasattr(salida1, 'total_seconds'):
                        total_seconds = int(salida1.total_seconds())
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        seconds = total_seconds % 60
                        cell.value = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                    else:
                        cell.value = str(salida1)
                else:
                    cell.value = ''
                cell.alignment = cell_alignment_center
                cell.border = border_style
                
                # Entrada 2
                cell = ws.cell(row=row_num, column=6)
                entrada2 = registro.get('Entrada 2')
                if entrada2:
                    if hasattr(entrada2, 'total_seconds'):
                        total_seconds = int(entrada2.total_seconds())
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        seconds = total_seconds % 60
                        cell.value = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                    else:
                        cell.value = str(entrada2)
                else:
                    cell.value = ''
                cell.alignment = cell_alignment_center
                cell.border = border_style
                
                # Salida 2
                cell = ws.cell(row=row_num, column=7)
                salida2 = registro.get('Salida 2')
                if salida2:
                    if hasattr(salida2, 'total_seconds'):
                        total_seconds = int(salida2.total_seconds())
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        seconds = total_seconds % 60
                        cell.value = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                    else:
                        cell.value = str(salida2)
                else:
                    cell.value = ''
                cell.alignment = cell_alignment_center
                cell.border = border_style
                
                # Estado
                cell = ws.cell(row=row_num, column=8)
                estado = registro.get('Estado', 'SIN MARCA')
                cell.value = estado
                cell.alignment = cell_alignment_center
                cell.border = border_style
                
                # Aplicar color segn estado
                if estado == 'ASISTIO':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100", bold=True)
                elif estado == 'TARDE':
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    cell.font = Font(color="9C6500", bold=True)
                elif estado == 'ASISTIO +5':
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006", bold=True)
                elif estado == 'SIN MARCA':
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    cell.font = Font(color="7F7F7F", bold=True)
                
                row_num += 1
            
            # ====================================================================
            # AJUSTAR ANCHO DE COLUMNAS
            # ====================================================================
            
            ws.column_dimensions['A'].width = 18  # Nmero Documento
            ws.column_dimensions['B'].width = 35  # Nombres Completos
            ws.column_dimensions['C'].width = 12  # Fecha
            ws.column_dimensions['D'].width = 12  # Entrada 1
            ws.column_dimensions['E'].width = 12  # Salida 1
            ws.column_dimensions['F'].width = 12  # Entrada 2
            ws.column_dimensions['G'].width = 12  # Salida 2
            ws.column_dimensions['H'].width = 15  # Estado
            
            # Fijar primera fila (encabezado)
            ws.freeze_panes = 'A2'
            
            # ====================================================================
            # GUARDAR EN MEMORIA Y ENVIAR
            # ====================================================================
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Nombre del archivo
            from datetime import datetime as dt
            timestamp = dt.now().strftime('%Y%m%d_%H%M%S')
            filename = f'Marcaciones_{fecha_desde}_a_{fecha_hasta}_{timestamp}.xlsx'
            
            print(f"[EXPORTAR_EXCEL] [OK] Archivo generado: {filename}")
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        
        except Error as e:
            print(f"[EXPORTAR_EXCEL] [X] Error SQL: {e}")
            return jsonify({'success': False, 'error': f'Error en la base de datos: {str(e)}'}), 500
    
    except Exception as e:
        print(f"[EXPORTAR_EXCEL] [X] Error general: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Error del servidor: {str(e)}'}), 500


# ============================================================================
# API: OBTENER MEMOS DE UN EMPLEADO
# ============================================================================

@marcacion_bp.route('/api/marcacion/memos-empleado', methods=['GET'])
@login_required
def obtener_memos_empleado():
    """Obtener lista de memos de amonestacin de un empleado"""
    try:
        num_documento = request.args.get('num_documento')
        
        if not num_documento:
            return jsonify({'success': False, 'error': 'Falta el nmero de documento'}), 400
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Error de conexin'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            print(f"[MEMOS_EMPLEADO] [...] Obteniendo memos para num_documento={num_documento}")
            
            # Llamar SP para obtener memos del empleado
            cursor.callproc('sp_ObtenerMemosEmpleado', (num_documento,))
            
            # Obtener resultados
            memos = []
            for result in cursor.stored_results():
                memos = result.fetchall()
            
            print(f"[MEMOS_EMPLEADO] [OK] {len(memos)} memos encontrados")
            
            # Serializar datos
            import datetime as dt_module
            for memo in memos:
                for key, value in memo.items():
                    if isinstance(value, datetime):
                        memo[key] = value.isoformat()
                    elif isinstance(value, dt_module.date):
                        memo[key] = value.strftime('%Y-%m-%d')
            
            cursor.close()
            connection.close()
            
            return jsonify({
                'success': True,
                'memos': memos
            }), 200
        
        except Error as e:
            print(f"[MEMOS_EMPLEADO] [X] Error SQL: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    except Exception as e:
        print(f"[MEMOS_EMPLEADO] [X] Error general: {e}")
        return jsonify({'success': False, 'error': 'Error del servidor'}), 500


# ============================================================================
# API: REPORTE DE CONTROL DE ASISTENCIA (NUEVO)
# ============================================================================

@marcacion_bp.route('/api/reportes/control-asistencia')
@login_required
def obtener_control_asistencia():
    """API: Obtener reporte de control de asistencia detallado por da usando SP"""
    from flask import current_app
    
    try:
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        num_usuario = request.args.get('num_usuario', None)
        
        if not fecha_inicio or not fecha_fin:
            return jsonify({'success': False, 'error': 'Se requieren fecha_inicio y fecha_fin'}), 400
        
        print(f"[CONTROL_ASISTENCIA] [-] Consultando: {fecha_inicio} al {fecha_fin}, usuario={num_usuario}")
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Error de conexin a BD'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            # Llamar al SP sp_reporte_asistencia_automatica
            if num_usuario and num_usuario != 'null' and num_usuario != '':
                cursor.callproc('sp_reporte_asistencia_automatica', [
                    fecha_inicio,
                    fecha_fin,
                    int(num_usuario)
                ])
            else:
                cursor.callproc('sp_reporte_asistencia_automatica', [
                    fecha_inicio,
                    fecha_fin,
                    None
                ])
            
            # Obtener resultados del SP
            registros = []
            for result in cursor.stored_results():
                registros = result.fetchall()
            
            cursor.close()
            connection.close()
            
            print(f"[CONTROL_ASISTENCIA] [OK] {len(registros)} registros obtenidos")
            
            return jsonify({
                'success': True,
                'data': registros
            }), 200
        
        except Error as e:
            print(f"[CONTROL_ASISTENCIA] [X] Error SQL: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    except Exception as e:
        print(f"[CONTROL_ASISTENCIA] [X] Error general: {e}")
        return jsonify({'success': False, 'error': 'Error del servidor'}), 500


# ============================================================================
# API: EXPORTAR CONTROL DE ASISTENCIA A EXCEL
# ============================================================================

@marcacion_bp.route('/api/reportes/control-asistencia/excel', methods=['GET'])
@login_required
def exportar_control_asistencia_excel():
    """Exportar reporte de control de asistencia a Excel"""
    try:
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return jsonify({'success': False, 'error': 'Se requieren fecha_inicio y fecha_fin'}), 400
        
        print(f"[EXPORTAR_CONTROL] [-] Exportando: {fecha_inicio} al {fecha_fin}")
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'error': 'Error de conexin a BD'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            # Llamar al SP sp_reporte_asistencia_automatica
            cursor.callproc('sp_reporte_asistencia_automatica', [
                fecha_inicio,
                fecha_fin,
                None
            ])
            
            # Obtener resultados del SP
            registros = []
            for result in cursor.stored_results():
                registros = result.fetchall()
            
            cursor.close()
            connection.close()
            
            print(f"[EXPORTAR_CONTROL] [OK] {len(registros)} registros obtenidos")
            
            if not registros:
                return jsonify({'success': False, 'error': 'No hay datos para exportar'}), 404
            
            # ====================================================================
            # CREAR ARCHIVO EXCEL
            # ====================================================================
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Control Asistencia"
            
            # Estilos
            header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=10)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            cell_alignment = Alignment(horizontal="left", vertical="center")
            cell_alignment_center = Alignment(horizontal="center", vertical="center")
            
            border_style = Border(
                left=Side(style='thin', color='D1D5DB'),
                right=Side(style='thin', color='D1D5DB'),
                top=Side(style='thin', color='D1D5DB'),
                bottom=Side(style='thin', color='D1D5DB')
            )
            
            # ====================================================================
            # ENCABEZADOS
            # ====================================================================
            
            headers = [
                'Empresa',
                'Nombres',
                'DNI/CE',
                'Cargo',
                'Sede',
                'Da',
                'Mes',
                'Ao',
                'H. Ingreso T1',
                'H. Salida T1',
                'Min T1',
                'Detalle T1',
                'H. Ingreso T2',
                'H. Salida T2',
                'Min T2',
                'Detalle T2'
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border_style
            
            # ====================================================================
            # DATOS
            # ====================================================================
            
            # Colores para detalles
            verde_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            verde_font = Font(color="006100", bold=True, size=10)
            amarillo_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            amarillo_font = Font(color="9C6500", bold=True, size=10)
            naranja_fill = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")
            naranja_font = Font(color="92400E", bold=True, size=10)
            gris_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            gris_font = Font(color="6B7280", size=10)
            
            row_num = 2
            for reg in registros:
                # Datos del empleado
                ws.cell(row=row_num, column=1, value=reg.get('EMPRESA', '')).alignment = cell_alignment_center
                ws.cell(row=row_num, column=2, value=reg.get('NOMBRES', '')).alignment = cell_alignment
                ws.cell(row=row_num, column=3, value=reg.get('DNI_CE', '')).alignment = cell_alignment_center
                ws.cell(row=row_num, column=4, value=reg.get('CARGO', '')).alignment = cell_alignment
                ws.cell(row=row_num, column=5, value=reg.get('SEDE_TRABAJO', '')).alignment = cell_alignment
                
                # Fecha
                ws.cell(row=row_num, column=6, value=reg.get('DIA', '')).alignment = cell_alignment_center
                ws.cell(row=row_num, column=7, value=reg.get('MES', '')).alignment = cell_alignment_center
                ws.cell(row=row_num, column=8, value=reg.get('ANO', '')).alignment = cell_alignment_center
                
                # Turno 1
                ws.cell(row=row_num, column=9, value=reg.get('H_INGRESO_T1', '')).alignment = cell_alignment_center
                ws.cell(row=row_num, column=10, value=reg.get('H_SALIDA_T1', '')).alignment = cell_alignment_center
                
                # Minutos T1 con formato horas:minutos
                min_t1 = reg.get('MINUTOS_T1', '')
                if min_t1 and min_t1 != '-' and min_t1 != 'Sin Marcacin':
                    try:
                        num_min = int(min_t1)
                        horas = abs(num_min) // 60
                        mins = abs(num_min) % 60
                        formato = f"{horas}:{mins:02d}"
                        if num_min > 0:
                            min_t1_display = f"+{formato}"
                        elif num_min < 0:
                            min_t1_display = f"-{formato}"
                        else:
                            min_t1_display = "0:00"
                    except:
                        min_t1_display = str(min_t1)
                else:
                    min_t1_display = '-'
                ws.cell(row=row_num, column=11, value=min_t1_display).alignment = cell_alignment_center
                
                # Detalle T1 con color
                cell_detalle_t1 = ws.cell(row=row_num, column=12, value=reg.get('DETALLE_T1', ''))
                cell_detalle_t1.alignment = cell_alignment_center
                if reg.get('DETALLE_T1') == 'ASISTENCIA':
                    cell_detalle_t1.fill = verde_fill
                    cell_detalle_t1.font = verde_font
                elif reg.get('DETALLE_T1') == 'TARDANZA':
                    cell_detalle_t1.fill = amarillo_fill
                    cell_detalle_t1.font = amarillo_font
                elif reg.get('DETALLE_T1') == 'SIN MARCACIN':
                    cell_detalle_t1.fill = gris_fill
                    cell_detalle_t1.font = gris_font
                
                # Turno 2
                ws.cell(row=row_num, column=13, value=reg.get('H_INGRESO_T2', '')).alignment = cell_alignment_center
                ws.cell(row=row_num, column=14, value=reg.get('H_SALIDA_T2', '')).alignment = cell_alignment_center
                
                # Minutos T2 con formato horas:minutos
                min_t2 = reg.get('MINUTOS_T2', '')
                if min_t2 and min_t2 != '-' and min_t2 != 'Sin Marcacin':
                    try:
                        num_min = int(min_t2)
                        horas = abs(num_min) // 60
                        mins = abs(num_min) % 60
                        formato = f"{horas}:{mins:02d}"
                        if num_min > 0:
                            min_t2_display = f"+{formato}"
                        elif num_min < 0:
                            min_t2_display = f"-{formato}"
                        else:
                            min_t2_display = "0:00"
                    except:
                        min_t2_display = str(min_t2)
                else:
                    min_t2_display = '-'
                ws.cell(row=row_num, column=15, value=min_t2_display).alignment = cell_alignment_center
                
                # Detalle T2 con color
                cell_detalle_t2 = ws.cell(row=row_num, column=16, value=reg.get('DETALLE_T2', ''))
                cell_detalle_t2.alignment = cell_alignment_center
                if reg.get('DETALLE_T2') == 'CUMPLI HORARIO':
                    cell_detalle_t2.fill = verde_fill
                    cell_detalle_t2.font = verde_font
                elif reg.get('DETALLE_T2') == 'SALIDA TEMPRANA':
                    cell_detalle_t2.fill = naranja_fill
                    cell_detalle_t2.font = naranja_font
                elif reg.get('DETALLE_T2') == 'SIN MARCACIN':
                    cell_detalle_t2.fill = gris_fill
                    cell_detalle_t2.font = gris_font
                
                # Aplicar bordes a toda la fila
                for col in range(1, 17):
                    ws.cell(row=row_num, column=col).border = border_style
                
                row_num += 1
            
            # ====================================================================
            # AJUSTAR ANCHO DE COLUMNAS
            # ====================================================================
            
            ws.column_dimensions['A'].width = 18  # Empresa
            ws.column_dimensions['B'].width = 35  # Nombres
            ws.column_dimensions['C'].width = 12  # DNI
            ws.column_dimensions['D'].width = 20  # Cargo
            ws.column_dimensions['E'].width = 15  # Sede
            ws.column_dimensions['F'].width = 8   # Da
            ws.column_dimensions['G'].width = 8   # Mes
            ws.column_dimensions['H'].width = 8   # Ao
            ws.column_dimensions['I'].width = 12  # H.Ingreso T1
            ws.column_dimensions['J'].width = 12  # H.Salida T1
            ws.column_dimensions['K'].width = 8   # Min T1
            ws.column_dimensions['L'].width = 15  # Detalle T1
            ws.column_dimensions['M'].width = 12  # H.Ingreso T2
            ws.column_dimensions['N'].width = 12  # H.Salida T2
            ws.column_dimensions['O'].width = 8   # Min T2
            ws.column_dimensions['P'].width = 15  # Detalle T2
            
            # Fijar primera fila (encabezado)
            ws.freeze_panes = 'A2'
            
            # ====================================================================
            # GUARDAR EN MEMORIA Y ENVIAR
            # ====================================================================
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Nombre del archivo
            from datetime import datetime as dt
            timestamp = dt.now().strftime('%Y%m%d_%H%M%S')
            filename = f'Control_Asistencia_{fecha_inicio}_a_{fecha_fin}_{timestamp}.xlsx'
            
            print(f"[EXPORTAR_CONTROL] [OK] Archivo generado: {filename}")
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        
        except Error as e:
            print(f"[EXPORTAR_CONTROL] [X] Error SQL: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    except Exception as e:
        print(f"[EXPORTAR_CONTROL] [X] Error general: {e}")
        return jsonify({'success': False, 'error': 'Error del servidor'}), 500


# ============================================================================
# API: OBTENER IDs DE MARCACIONES POR DOCUMENTO Y FECHA
# ============================================================================

@marcacion_bp.route('/api/marcacion/obtener-ids', methods=['GET'])
@login_required
def obtener_ids_marcacion():
    """Obtener IDs de las marcaciones de un empleado en una fecha específica"""
    try:
        num_documento_raw = request.args.get('num_documento')
        fecha = request.args.get('fecha')
        
        print(f"[OBTENER_IDS] [+] Parámetros recibidos: num_documento_raw={num_documento_raw} (tipo={type(num_documento_raw).__name__}), fecha={fecha}")
        
        if not num_documento_raw or not fecha:
            print("[OBTENER_IDS] [X] Faltan parámetros")
            return jsonify({'success': False, 'error': 'Se requieren num_documento y fecha'}), 400
        
        # num_documento_raw es el DNI (documento_numero) como string
        num_documento_str = str(num_documento_raw).strip()
        
        print(f"[OBTENER_IDS] Buscando marcaciones: DNI={num_documento_str}, Fecha={fecha}")
        
        connection = get_db_connection()
        if not connection:
            print("[OBTENER_IDS] [X] Error de conexión a BD")
            return jsonify({'success': False, 'error': 'Error de conexión a BD'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            print(f"[OBTENER_IDS] Llamando SP sp_ObtenerIdsMarcacion('{num_documento_str}', '{fecha}')")
            cursor.callproc('sp_ObtenerIdsMarcacion', [num_documento_str, fecha])
            
            # Obtener resultados
            ids = None
            for result in cursor.stored_results():
                ids = result.fetchone()
            
            cursor.close()
            connection.close()
            
            if ids:
                print(f"[OBTENER_IDS] [OK] IDs encontrados: T1E={ids['id_t1_entrada']}, T1S={ids['id_t1_salida']}, T2E={ids['id_t2_entrada']}, T2S={ids['id_t2_salida']}")
                return jsonify({
                    'success': True,
                    'data': {
                        'id_t1_entrada': ids['id_t1_entrada'],
                        'id_t1_salida': ids['id_t1_salida'],
                        'id_t2_entrada': ids['id_t2_entrada'],
                        'id_t2_salida': ids['id_t2_salida']
                    }
                })
            else:
                print("[OBTENER_IDS] [!] SP no retornó resultados")
                return jsonify({'success': False, 'error': 'No se encontraron marcaciones'}), 404
        
        except Error as e:
            print(f"[OBTENER_IDS] [X] Error SQL: {e}")
            import traceback
            traceback.print_exc()
            if connection:
                connection.close()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    except Exception as e:
        print(f"[OBTENER_IDS] [X] Error general: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': 'Error del servidor'}), 500


# ============================================================================
# API: EDITAR MARCACION (HORAS)
# ============================================================================

@marcacion_bp.route('/api/marcacion/editar', methods=['PUT'])
@login_required
def editar_marcacion():
    """Editar hora de entrada/salida de una marcación"""
    try:
        data = request.get_json()
        
        print(f"[EDITAR_MARCACION] [+] Datos recibidos: {data}")
        
        if not data:
            print("[EDITAR_MARCACION] [X] No se recibieron datos")
            return jsonify({'success': False, 'error': 'Datos no proporcionados'}), 400
        
        id_marcacion_raw = data.get('id_marcacion')
        nueva_hora = data.get('nueva_hora')  # Formato: "HH:MM"
        fecha_original = data.get('fecha_original')  # Formato: "YYYY-MM-DD"
        
        print(f"[EDITAR_MARCACION] Parámetros: id_marcacion={id_marcacion_raw} (tipo={type(id_marcacion_raw).__name__}), nueva_hora={nueva_hora}, fecha_original={fecha_original}")
        
        if not id_marcacion_raw or not nueva_hora or not fecha_original:
            print(f"[EDITAR_MARCACION] [X] Faltan parámetros")
            return jsonify({'success': False, 'error': 'Faltan parámetros: id_marcacion, nueva_hora, fecha_original'}), 400
        
        # Convertir id_marcacion a int
        try:
            id_marcacion = int(id_marcacion_raw)
        except (ValueError, TypeError):
            print(f"[EDITAR_MARCACION] [X] id_marcacion no es numérico: {id_marcacion_raw}")
            return jsonify({'success': False, 'error': f'id_marcacion inválido: {id_marcacion_raw}'}), 400
        
        print(f"[EDITAR_MARCACION] Editando marcación ID={id_marcacion}: fecha={fecha_original}, hora={nueva_hora}")
        
        connection = get_db_connection()
        if not connection:
            print("[EDITAR_MARCACION] [X] Error de conexión a BD")
            return jsonify({'success': False, 'error': 'Error de conexión a BD'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            # Llamar al SP sp_EditarMarcacion
            print(f"[EDITAR_MARCACION] Llamando SP sp_EditarMarcacion({id_marcacion}, {nueva_hora}, {fecha_original})")
            cursor.execute("CALL sp_EditarMarcacion(%s, %s, %s, @p_mensaje)", (id_marcacion, nueva_hora, fecha_original))
            
            # Consumir result sets pendientes del SP
            while cursor.nextset():
                pass
            
            # Obtener parámetro OUT del SP
            cursor.execute("SELECT @p_mensaje AS mensaje")
            out_result = cursor.fetchone()
            mensaje = out_result['mensaje'] if out_result and out_result.get('mensaje') else 'Actualización realizada'
            
            connection.commit()
            cursor.close()
            connection.close()
            
            print(f"[EDITAR_MARCACION] [OK] {mensaje}")
            
            return jsonify({
                'success': True,
                'message': mensaje,
                'data': {
                    'id_marcacion': id_marcacion,
                    'nueva_hora': nueva_hora,
                    'fecha_original': fecha_original
                }
            })
        
        except Error as e:
            print(f"[EDITAR_MARCACION] [X] Error SQL: {e}")
            import traceback
            traceback.print_exc()
            if connection:
                connection.close()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    except Exception as e:
        print(f"[EDITAR_MARCACION] [X] Error general: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': 'Error del servidor'}), 500
